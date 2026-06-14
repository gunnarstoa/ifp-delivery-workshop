"""IFP Delivery Workshop — auth gateway and static-page server.

Serves the existing docs/, css/, data/ trees behind username/password auth.
Logs every authenticated request to page_views for later analysis.
Facilitator-only pages additionally require users.is_facilitator = 1.
"""
import calendar as calmod
import json
import os
import re
import secrets
import sqlite3
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import bcrypt
from flask import (
    Flask,
    abort,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_from_directory,
    session,
    url_for,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = Path(os.environ.get("WORKSHOP_DB", REPO_ROOT / "workshop.db"))
SECRET_KEY = os.environ.get("WORKSHOP_SECRET_KEY") or secrets.token_hex(32)

# Paths that require an additional is_facilitator check.
FACILITATOR_PATH_RE = re.compile(
    r"^/(docs/(15-facilitator|16-facilitator-toolkit|cohort-tracker)|data/|admin|w/[a-z0-9-]+/15-facilitator|w/[a-z0-9-]+/16-facilitator-toolkit|w/[a-z0-9-]+/cohort-tracker)"
)
# Paths that bypass auth entirely (login form, static assets, health).
PUBLIC_PREFIXES = ("/login", "/static/", "/css/", "/js/", "/img/", "/assets/", "/favicon", "/healthz", "/robots.txt")

app = Flask(__name__, static_folder=None)
app.config.update(
    SECRET_KEY=SECRET_KEY,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("WORKSHOP_PRODUCTION") == "1",
    PERMANENT_SESSION_LIFETIME=60 * 60 * 8,  # 8 hours
)


# ---------------------------- DB helpers ----------------------------

def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(str(DB_PATH))
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
        g.db.execute("PRAGMA journal_mode = WAL")
    return g.db


@app.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    """Create tables if missing. Called at startup and from `flask init-db`."""
    conn = sqlite3.connect(str(DB_PATH))
    try:
        with open(REPO_ROOT / "backend" / "schema.sql") as f:
            conn.executescript(f.read())
        # Idempotent column adds for migrations beyond the initial deploy.
        def cols(t):
            return {r[1] for r in conn.execute(f"PRAGMA table_info({t})").fetchall()}
        if "passing_score" not in cols("survey_templates"):
            conn.execute("ALTER TABLE survey_templates ADD COLUMN passing_score INTEGER NOT NULL DEFAULT 70")
        if "max_attempts" not in cols("survey_templates"):
            conn.execute("ALTER TABLE survey_templates ADD COLUMN max_attempts INTEGER")
        if "questions_per_attempt" not in cols("survey_templates"):
            conn.execute("ALTER TABLE survey_templates ADD COLUMN questions_per_attempt INTEGER")
        if "shuffle_options" not in cols("survey_templates"):
            conn.execute("ALTER TABLE survey_templates ADD COLUMN shuffle_options INTEGER NOT NULL DEFAULT 1")
        if "score" not in cols("survey_responses"):
            conn.execute("ALTER TABLE survey_responses ADD COLUMN score INTEGER")
        if "passed" not in cols("survey_responses"):
            conn.execute("ALTER TABLE survey_responses ADD COLUMN passed INTEGER")
        if "attempt_number" not in cols("survey_responses"):
            conn.execute("ALTER TABLE survey_responses ADD COLUMN attempt_number INTEGER")
        if "submitted_at" not in cols("survey_responses"):
            conn.execute("ALTER TABLE survey_responses ADD COLUMN submitted_at TIMESTAMP")
        conn.commit()
    finally:
        conn.close()


# ---------------------------- Auth helpers ----------------------------

def current_user():
    uid = session.get("uid")
    if uid is None:
        return None
    if "user" in g:
        return g.user
    row = get_db().execute(
        "SELECT id, username, display_name, email, is_facilitator, is_active FROM users WHERE id = ?",
        (uid,),
    ).fetchone()
    if row is None or not row["is_active"]:
        session.clear()
        return None
    g.user = row
    return row


def is_facilitator_path(path: str) -> bool:
    return bool(FACILITATOR_PATH_RE.match(path))


def is_public_path(path: str) -> bool:
    return any(path.startswith(p) for p in PUBLIC_PREFIXES)


def login_required(path: str):
    """Raise abort(401)/abort(403) if request lacks the right auth for `path`."""
    if is_public_path(path):
        return None
    user = current_user()
    if user is None:
        return redirect(url_for("login", next=path))
    if is_facilitator_path(path) and not user["is_facilitator"]:
        abort(403)
    return None


@app.before_request
def auth_gate():
    path = request.path
    redirect_or_none = login_required(path)
    if redirect_or_none is not None:
        return redirect_or_none


@app.after_request
def log_page_view(resp):
    # Only log authenticated GETs of HTML/JSON content we actually served.
    if request.method != "GET":
        return resp
    if is_public_path(request.path):
        return resp
    user = g.get("user")
    if user is None:
        return resp
    if resp.status_code >= 400 and resp.status_code != 304:
        # Still log denied access so facilitator can review
        pass
    try:
        get_db().execute(
            "INSERT INTO page_views (user_id, path, query, status, referer, user_agent) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                user["id"],
                request.path,
                request.query_string.decode("utf-8", errors="replace") or None,
                resp.status_code,
                request.referrer,
                request.user_agent.string[:500] if request.user_agent else None,
            ),
        )
        get_db().commit()
    except sqlite3.Error:
        pass
    return resp


# ---------------------------- Routes: login ----------------------------

@app.route("/")
def home():
    user = current_user()
    if user is None:
        return redirect(url_for("login"))
    if user["is_facilitator"]:
        return redirect(url_for("admin_dashboard"))
    # Send participant to the workshop of their most recent enrollment.
    row = get_db().execute(
        "SELECT w.slug FROM session_participants sp "
        "JOIN sessions s ON s.id = sp.session_id "
        "JOIN workshops w ON w.id = s.workshop_id "
        "WHERE sp.user_id = ? AND sp.removed_at IS NULL AND s.status != 'archived' "
        "ORDER BY sp.enrolled_at DESC LIMIT 1",
        (user["id"],),
    ).fetchone()
    slug = row["slug"] if row else "ifp"
    return redirect(f"/w/{slug}/01-overview.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        ok = False
        row = get_db().execute(
            "SELECT id, username, password_hash, is_active FROM users WHERE username = ?",
            (username,),
        ).fetchone()
        if row and row["is_active"]:
            try:
                ok = bcrypt.checkpw(password.encode("utf-8"), row["password_hash"].encode("utf-8"))
            except ValueError:
                ok = False
        get_db().execute(
            "INSERT INTO login_events (username_attempted, succeeded, ip, user_agent) VALUES (?, ?, ?, ?)",
            (username, 1 if ok else 0, request.headers.get("X-Forwarded-For", request.remote_addr), request.user_agent.string[:500] if request.user_agent else None),
        )
        get_db().commit()
        if ok:
            session.clear()
            session["uid"] = row["id"]
            session.permanent = True
            get_db().execute("UPDATE users SET last_login_at = CURRENT_TIMESTAMP WHERE id = ?", (row["id"],))
            get_db().commit()
            nxt = request.args.get("next") or request.form.get("next") or _default_landing_for_user(row["id"]) or "/"
            if not nxt.startswith("/") or nxt.startswith("//"):
                nxt = "/"
            return redirect(nxt)
        time.sleep(0.5)  # rate-limit failed attempts modestly
        return render_template("login.html", error="Invalid username or password.", next=request.args.get("next") or "")
    return render_template("login.html", error=None, next=request.args.get("next") or "")


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------- Routes: static page serving ----------------------------

def _serve_repo_file(subdir: str, filename: str):
    base = (REPO_ROOT / subdir).resolve()
    target = (base / filename).resolve()
    if not str(target).startswith(str(base)):
        abort(404)
    if not target.is_file():
        abort(404)
    return send_from_directory(str(base), filename)


@app.route("/docs/<path:filename>")
def serve_docs(filename):
    # Backward-compat: if a file has been moved to docs/ifp/, 301 to the new path.
    ifp_path = (REPO_ROOT / "docs" / "ifp" / filename).resolve()
    if ifp_path.is_file() and str(ifp_path).startswith(str((REPO_ROOT / "docs" / "ifp").resolve())):
        return redirect(f"/w/ifp/{filename}", code=301)
    return _serve_repo_file("docs", filename)


@app.route("/w/<workshop_slug>/<path:filename>")
def serve_workshop(workshop_slug, filename):
    # Defense-in-depth: workshop_slug must match SLUG_RE shape; route param itself doesn't enforce it.
    if not re.fullmatch(r"[a-z0-9-]{1,40}", workshop_slug):
        abort(404)
    return _serve_repo_file(f"docs/{workshop_slug}", filename)


@app.route("/data/<path:filename>")
def serve_data(filename):
    return _serve_repo_file("data", filename)


@app.route("/css/<path:filename>")
def serve_css(filename):
    return _serve_repo_file("css", filename)


@app.route("/js/<path:filename>")
def serve_js(filename):
    return _serve_repo_file("js", filename)


@app.route("/img/<path:filename>")
def serve_img(filename):
    return _serve_repo_file("img", filename)


@app.route("/assets/<path:filename>")
def serve_assets(filename):
    return _serve_repo_file("assets", filename)


@app.route("/favicon.ico")
def favicon():
    fav = REPO_ROOT / "favicon.ico"
    if fav.is_file():
        return send_from_directory(str(REPO_ROOT), "favicon.ico")
    abort(404)


@app.route("/healthz")
def healthz():
    return "ok", 200


# ---------------------------- Routes: admin ----------------------------

def _require_facilitator():
    user = current_user()
    if user is None:
        return None, redirect(url_for("login", next=request.path))
    if not user["is_facilitator"]:
        abort(403)
    return user, None


CAL_COLORS = ["blue", "green", "purple", "orange", "teal", "pink", "amber", "indigo"]


def _parse_month_param(raw):
    """Parse a ?month=YYYY-MM query param; fall back to current month."""
    if raw:
        m = re.fullmatch(r"(\d{4})-(\d{2})", raw)
        if m:
            y, mo = int(m.group(1)), int(m.group(2))
            if 1 <= mo <= 12 and 2000 <= y <= 2100:
                return y, mo
    today = date.today()
    return today.year, today.month


def _month_grid(year, month, sessions_with_workshops):
    """Build a list of weeks; each week is a list of day dicts. Week starts Sunday."""
    today = date.today()
    cal = calmod.Calendar(firstweekday=calmod.SUNDAY)
    weeks = []
    for week in cal.monthdatescalendar(year, month):
        wrow = []
        for d in week:
            day_sessions = []
            for s in sessions_with_workshops:
                if s["_start"] <= d <= s["_end"]:
                    day_sessions.append({
                        "slug": s["slug"],
                        "name": s["name"],
                        "workshop_slug": s["workshop_slug"],
                        "workshop_name": s["workshop_name"],
                        "color": CAL_COLORS[(s["workshop_id"] - 1) % len(CAL_COLORS)],
                        "is_start": d == s["_start"],
                        "is_end": d == s["_end"],
                    })
            wrow.append({
                "date": d,
                "in_month": d.month == month,
                "is_today": d == today,
                "sessions": day_sessions,
            })
        weeks.append(wrow)
    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month = (date(year, month, 28) + timedelta(days=10)).replace(day=1)
    return {
        "weeks": weeks,
        "year": year,
        "month": month,
        "month_name": calmod.month_name[month],
        "prev_param": f"{prev_month.year:04d}-{prev_month.month:02d}",
        "next_param": f"{next_month.year:04d}-{next_month.month:02d}",
        "today_param": f"{today.year:04d}-{today.month:02d}",
    }


def _sessions_overlapping_month(year, month):
    """All non-archived sessions that touch the given month, joined to workshop info."""
    first = date(year, month, 1)
    last_day = calmod.monthrange(year, month)[1]
    last = date(year, month, last_day)
    rows = get_db().execute(
        "SELECT s.id, s.slug, s.name, s.start_date, s.end_date, s.status, "
        "       w.id AS workshop_id, w.slug AS workshop_slug, w.name AS workshop_name "
        "FROM sessions s JOIN workshops w ON w.id = s.workshop_id "
        "WHERE s.status != 'archived' AND w.status != 'archived' "
        "  AND s.end_date >= ? AND s.start_date <= ? "
        "ORDER BY s.start_date ASC",
        (first.isoformat(), last.isoformat()),
    ).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["_start"] = date.fromisoformat(d["start_date"])
        d["_end"] = date.fromisoformat(d["end_date"])
        out.append(d)
    return out


@app.route("/admin")
def admin_dashboard():
    user, redir = _require_facilitator()
    if redir:
        return redir
    workshops = get_db().execute(
        "SELECT w.id, w.slug, w.name, w.short_description, w.contact_email, w.status, w.created_at, "
        "       u.username AS created_by_username, "
        "       (SELECT COUNT(*) FROM sessions s WHERE s.workshop_id = w.id AND s.status != 'archived') AS session_count "
        "FROM workshops w LEFT JOIN users u ON u.id = w.created_by "
        "WHERE w.status != 'archived' "
        "ORDER BY w.created_at DESC"
    ).fetchall()
    # Color per workshop for the legend; matches CAL_COLORS cycle
    workshops_for_legend = [
        {"slug": w["slug"], "name": w["name"], "color": CAL_COLORS[(w["id"] - 1) % len(CAL_COLORS)]}
        for w in workshops
    ]
    year, month = _parse_month_param(request.args.get("month"))
    sessions = _sessions_overlapping_month(year, month)
    grid = _month_grid(year, month, sessions)
    return render_template(
        "admin_dashboard.html",
        user=user,
        workshops=workshops,
        legend=workshops_for_legend,
        grid=grid,
    )


SLUG_RE = re.compile(r"^[a-z0-9]([a-z0-9-]{1,38}[a-z0-9])?$")


@app.route("/admin/workshops/new", methods=["GET", "POST"])
def admin_workshop_new():
    user, redir = _require_facilitator()
    if redir:
        return redir
    errors = {}
    form = {"slug": "", "name": "", "short_description": "", "contact_email": ""}
    if request.method == "POST":
        form["slug"] = (request.form.get("slug") or "").strip().lower()
        form["name"] = (request.form.get("name") or "").strip()
        form["short_description"] = (request.form.get("short_description") or "").strip()
        form["contact_email"] = (request.form.get("contact_email") or "").strip()
        if not SLUG_RE.fullmatch(form["slug"]):
            errors["slug"] = "3-40 chars, lowercase letters/digits/hyphen, no leading/trailing hyphen"
        if not form["name"]:
            errors["name"] = "required"
        if not errors:
            try:
                cur = get_db().execute(
                    "INSERT INTO workshops (slug, name, short_description, contact_email, created_by) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (form["slug"], form["name"], form["short_description"] or None, form["contact_email"] or None, user["id"]),
                )
                get_db().commit()
                return redirect(url_for("admin_workshop_detail", slug=form["slug"]))
            except sqlite3.IntegrityError:
                errors["slug"] = f"workshop '{form['slug']}' already exists"
    return render_template("admin_workshop_new.html", user=user, form=form, errors=errors)


TOOLKIT_ROOT = REPO_ROOT / "toolkit"


def _parse_toolkit_file(path: Path):
    """Parse a toolkit markdown file: first '# Title' line, optional '> description',
    rest is the copyable body. Returns dict or None if file is missing/empty."""
    try:
        text = path.read_text(encoding="utf-8")
    except (OSError, UnicodeError):
        return None
    lines = text.splitlines()
    title = path.stem
    description = ""
    i = 0
    while i < len(lines) and not lines[i].strip():
        i += 1
    if i < len(lines) and lines[i].startswith("# "):
        title = lines[i][2:].strip()
        i += 1
    if i < len(lines) and lines[i].startswith("> "):
        description = lines[i][2:].strip()
        i += 1
    while i < len(lines) and not lines[i].strip():
        i += 1
    body = "\n".join(lines[i:]).rstrip()
    return {
        "filename": path.name,
        "slug": path.stem,
        "title": title,
        "description": description,
        "body": body,
        "size_kb": round(path.stat().st_size / 1024, 1),
        "modified": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }


def _list_toolkit_assets(workshop_slug: str):
    base = (TOOLKIT_ROOT / workshop_slug).resolve()
    if not base.is_dir() or not str(base).startswith(str(TOOLKIT_ROOT.resolve())):
        return None
    assets = []
    for p in sorted(base.glob("*.md")):
        a = _parse_toolkit_file(p)
        if a:
            assets.append(a)
    return assets


def _load_toolkit_asset(workshop_slug: str, asset_slug: str):
    base = (TOOLKIT_ROOT / workshop_slug).resolve()
    if not base.is_dir() or not str(base).startswith(str(TOOLKIT_ROOT.resolve())):
        return None
    if not re.fullmatch(r"[a-z0-9._-]{1,80}", asset_slug):
        return None
    path = (base / f"{asset_slug}.md").resolve()
    if not path.is_file() or not str(path).startswith(str(base)):
        return None
    return _parse_toolkit_file(path)


def _likert_avg_across_responses(responses):
    """Mean Likert value across every Likert-classifiable cell of every response."""
    nums = []
    for r in responses:
        for k, v in r["responses"].items():
            if _is_metadata_header(k) or v is None:
                continue
            n = LIKERT_MAP.get(str(v).strip().lower())
            if n is not None:
                nums.append(n)
    return (round(sum(nums) / len(nums), 2), len(nums)) if nums else (None, 0)


def _kc_pass_rate(responses):
    """Pass-rate from a knowledge-check survey: scan any pass/fail-shaped column."""
    if not responses:
        return None, 0
    total = passed = 0
    for r in responses:
        for k, v in r["responses"].items():
            if _is_metadata_header(k) or v is None:
                continue
            s = str(v).strip()
            if PASS_FAIL_RE.match(s):
                total += 1
                passed += 1
            elif FAIL_RE.match(s):
                total += 1
    if total == 0:
        return None, 0
    return round(passed / total * 100, 1), total


def _workshop_summary(workshop_id):
    """Aggregate metrics across all non-archived sessions for a workshop."""
    from collections import Counter
    db = get_db()
    sessions = db.execute(
        "SELECT id, slug, name, start_date, end_date "
        "FROM sessions WHERE workshop_id = ? AND status != 'archived' "
        "ORDER BY start_date",
        (workshop_id,),
    ).fetchall()
    if not sessions:
        return None
    session_ids = [s["id"] for s in sessions]
    ph = ",".join("?" * len(session_ids))
    participants = db.execute(
        f"SELECT sp.partner, sp.excluded, sp.session_id "
        f"FROM session_participants sp "
        f"WHERE sp.session_id IN ({ph}) AND sp.removed_at IS NULL",
        session_ids,
    ).fetchall()
    counted = [p for p in participants if not p["excluded"]]
    partner_counts = Counter(p["partner"] or "Unknown" for p in counted)
    top_partners = partner_counts.most_common(10)
    by_session_participants = Counter(p["session_id"] for p in counted)

    pre_nums, post_nums = [], []
    session_summaries = []
    for s in sessions:
        surveys = _load_session_surveys(s["id"])
        pre = surveys.get("pre")
        post = surveys.get("post")
        kc = surveys.get("knowledge")
        pre_avg, pre_n = (None, 0)
        post_avg, post_n = (None, 0)
        kc_rate, kc_total = (None, 0)
        if pre:
            pre_resp = _load_survey_responses(pre["id"])
            pre_avg, pre_n = _likert_avg_across_responses(pre_resp)
            if pre_avg is not None:
                pre_nums.extend([pre_avg] * pre_n)
        if post:
            post_resp = _load_survey_responses(post["id"])
            post_avg, post_n = _likert_avg_across_responses(post_resp)
            if post_avg is not None:
                post_nums.extend([post_avg] * post_n)
        if kc:
            kc_resp = _load_survey_responses(kc["id"])
            kc_rate, kc_total = _kc_pass_rate(kc_resp)
        delta = round(post_avg - pre_avg, 2) if (pre_avg is not None and post_avg is not None) else None
        session_summaries.append({
            "slug": s["slug"], "name": s["name"],
            "start_date": s["start_date"], "end_date": s["end_date"],
            "participants": by_session_participants.get(s["id"], 0),
            "pre_avg": pre_avg, "post_avg": post_avg, "delta": delta,
            "pre_n": pre_n, "post_n": post_n,
            "kc_pass_rate": kc_rate, "kc_n": kc_total,
        })

    workshop_pre_avg = round(sum(pre_nums) / len(pre_nums), 2) if pre_nums else None
    workshop_post_avg = round(sum(post_nums) / len(post_nums), 2) if post_nums else None
    workshop_delta = (round(workshop_post_avg - workshop_pre_avg, 2)
                      if workshop_pre_avg is not None and workshop_post_avg is not None else None)

    return {
        "session_count": len(sessions),
        "total_counted": len(counted),
        "total_partners": len(partner_counts),
        "first_date": sessions[0]["start_date"],
        "last_date": sessions[-1]["end_date"],
        "top_partners": top_partners,
        "session_summaries": session_summaries,
        "workshop_pre_avg": workshop_pre_avg,
        "workshop_post_avg": workshop_post_avg,
        "workshop_delta": workshop_delta,
    }


def _list_workshop_pages(workshop_slug):
    """Scan docs/<workshop_slug>/ for .html pages. Returns sorted list with mtimes."""
    base = (REPO_ROOT / "docs" / workshop_slug).resolve()
    if not base.is_dir():
        return None
    pages = []
    for p in sorted(base.glob("*.html")):
        st = p.stat()
        pages.append({
            "filename": p.name,
            "size_kb": round(st.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            "url": f"/w/{workshop_slug}/{p.name}",
        })
    return pages


@app.route("/admin/workshops/<slug>")
def admin_workshop_detail(slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    workshop = get_db().execute(
        "SELECT w.*, u.username AS created_by_username "
        "FROM workshops w LEFT JOIN users u ON u.id = w.created_by "
        "WHERE w.slug = ?",
        (slug,),
    ).fetchone()
    if workshop is None:
        abort(404)
    sessions = get_db().execute(
        "SELECT s.slug, s.name, s.start_date, s.end_date, s.status, s.created_at "
        "FROM sessions s WHERE s.workshop_id = ? AND s.status != 'archived' "
        "ORDER BY s.start_date DESC",
        (workshop["id"],),
    ).fetchall()
    pages = _list_workshop_pages(slug)
    content_root = f"docs/{slug}/"
    toolkit_assets = _list_toolkit_assets(slug)
    toolkit_root = f"toolkit/{slug}/"
    summary = _workshop_summary(workshop["id"])
    return render_template(
        "admin_workshop_detail.html",
        user=user,
        workshop=workshop,
        sessions=sessions,
        pages=pages,
        content_root=content_root,
        toolkit_assets=toolkit_assets,
        toolkit_root=toolkit_root,
        summary=summary,
    )


@app.route("/admin/workshops/<slug>/toolkit")
def admin_workshop_toolkit(slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    workshop = get_db().execute("SELECT * FROM workshops WHERE slug = ?", (slug,)).fetchone()
    if workshop is None:
        abort(404)
    assets = _list_toolkit_assets(slug)
    return render_template(
        "admin_workshop_toolkit.html",
        user=user, workshop=workshop, assets=assets,
        toolkit_root=f"toolkit/{slug}/",
    )


@app.route("/admin/workshops/<slug>/toolkit/<asset_slug>")
def admin_workshop_toolkit_asset(slug, asset_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    workshop = get_db().execute("SELECT * FROM workshops WHERE slug = ?", (slug,)).fetchone()
    if workshop is None:
        abort(404)
    asset = _load_toolkit_asset(slug, asset_slug)
    if asset is None:
        abort(404)
    return render_template(
        "admin_workshop_toolkit_asset.html",
        user=user, workshop=workshop, asset=asset,
    )


@app.route("/admin/workshops/<slug>/sessions/new", methods=["GET", "POST"])
def admin_session_new(slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    workshop = get_db().execute("SELECT * FROM workshops WHERE slug = ?", (slug,)).fetchone()
    if workshop is None:
        abort(404)
    errors = {}
    today = date.today()
    default_end = (today + timedelta(days=4)).isoformat()
    form = {
        "slug": request.form.get("slug", "").strip().lower() if request.method == "POST" else today.isoformat(),
        "name": request.form.get("name", "").strip() if request.method == "POST" else f"{today.strftime('%B %-d, %Y')} Session",
        "start_date": request.form.get("start_date", today.isoformat()) if request.method == "POST" else today.isoformat(),
        "end_date": request.form.get("end_date", default_end) if request.method == "POST" else default_end,
        "notes": request.form.get("notes", "").strip() if request.method == "POST" else "",
    }
    if request.method == "POST":
        if not SLUG_RE.fullmatch(form["slug"]):
            errors["slug"] = "3-40 chars, lowercase letters/digits/hyphen, no leading/trailing hyphen"
        if not form["name"]:
            errors["name"] = "required"
        try:
            sd = date.fromisoformat(form["start_date"])
            ed = date.fromisoformat(form["end_date"])
            if ed < sd:
                errors["end_date"] = "end date must be on or after start date"
        except ValueError:
            errors["start_date"] = "invalid date"
        if not errors:
            try:
                get_db().execute(
                    "INSERT INTO sessions (workshop_id, slug, name, start_date, end_date, notes, created_by) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (workshop["id"], form["slug"], form["name"], form["start_date"], form["end_date"], form["notes"] or None, user["id"]),
                )
                get_db().commit()
                return redirect(url_for("admin_session_detail", slug=slug, session_slug=form["slug"]))
            except sqlite3.IntegrityError:
                errors["slug"] = f"a session with slug '{form['slug']}' already exists in this workshop"
    return render_template("admin_session_new.html", user=user, workshop=workshop, form=form, errors=errors)


@app.route("/admin/workshops/<slug>/sessions/<session_slug>")
def admin_session_detail(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    row = get_db().execute(
        "SELECT s.*, w.slug AS workshop_slug, w.name AS workshop_name, u.username AS created_by_username "
        "FROM sessions s JOIN workshops w ON w.id = s.workshop_id "
        "LEFT JOIN users u ON u.id = s.created_by "
        "WHERE w.slug = ? AND s.slug = ?",
        (slug, session_slug),
    ).fetchone()
    if row is None:
        abort(404)
    counts = get_db().execute(
        "SELECT COUNT(*) AS total, SUM(CASE WHEN excluded = 0 THEN 1 ELSE 0 END) AS included "
        "FROM session_participants WHERE session_id = ? AND removed_at IS NULL",
        (row["id"],),
    ).fetchone()
    monitors = _session_monitors(row["id"])
    return render_template(
        "admin_session_detail.html",
        user=user, session=row,
        participant_total=counts["total"] or 0,
        participant_included=counts["included"] or 0,
        monitors=monitors,
    )


EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PARTNERS_FILE = REPO_ROOT / "data" / "cohorts" / "_partners.json"

LIKERT_MAP = {
    "very comfortable": 5,
    "comfortable": 4,
    "somewhat comfortable": 4,
    "neither comfortable nor uncomfortable": 3,
    "neutral": 3,
    "somewhat uncomfortable": 2,
    "uncomfortable": 2,
    "very uncomfortable": 1,
    "not at all comfortable": 1,
    "strongly agree": 5,
    "agree": 4,
    "somewhat agree": 4,
    "neither agree nor disagree": 3,
    "somewhat disagree": 2,
    "disagree": 2,
    "strongly disagree": 1,
}

LIKERT_ORDER = ["Very comfortable", "Somewhat comfortable", "Neither comfortable nor uncomfortable", "Somewhat uncomfortable", "Very uncomfortable"]

CATEGORICAL_ALIASES = {
    "master anaplanner": "Certified Master Anaplanner",
}

SURVEY_METADATA_HEADERS = {"id", "start time", "completion time", "email", "name", "last modified time"}
PASS_FAIL_RE = re.compile(r"^(pass|passed|yes|y|true|1)$", re.IGNORECASE)
FAIL_RE = re.compile(r"^(fail|failed|no|n|false|0)$", re.IGNORECASE)


def _load_partners():
    """Load partner directory and build a domain→partner_name lookup. Cached on app.config."""
    cached = app.config.get("partners_cached")
    if cached is not None:
        return cached
    try:
        with open(PARTNERS_FILE) as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        data = {"partners": [], "excludePartners": []}
    domains = {}
    for p in data.get("partners", []):
        for d in p.get("domains", []):
            domains[d.strip().lower()] = p["name"]
    out = {"domains": domains, "exclude": set(data.get("excludePartners", []))}
    app.config["partners_cached"] = out
    return out


def _lookup_partner(email):
    """Return (partner_name|None, excluded_bool) for an email address."""
    p = _load_partners()
    domain = email.rsplit("@", 1)[-1].lower() if "@" in email else ""
    name = p["domains"].get(domain)
    if name is None:
        return None, False
    return name, name in p["exclude"]


def _parse_email_list(text):
    """Pull unique lowercased emails out of pasted text. Handles 'Name <email>'."""
    seen = set()
    out = []
    for line in (text or "").splitlines():
        m = EMAIL_RE.search(line)
        if not m:
            continue
        email = m.group(0).lower()
        if email in seen:
            continue
        seen.add(email)
        out.append(email)
    return out


def _add_participants_to_session(session_id, emails, enrolled_by_user_id):
    """For each email: create user if missing, enroll into session.
    Returns list of dicts with username, partner, status, initial_password (or None)."""
    db = get_db()
    results = []
    any_reactivated = False
    for email in emails:
        partner_name, excluded = _lookup_partner(email)
        existing = db.execute("SELECT id FROM users WHERE username = ?", (email,)).fetchone()
        if existing:
            user_id = existing["id"]
            initial_password = None
            user_status = "existing"
        else:
            initial_password = secrets.token_urlsafe(10)
            pw_hash = bcrypt.hashpw(initial_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
            cur = db.execute(
                "INSERT INTO users (username, email, password_hash, is_facilitator, is_active) "
                "VALUES (?, ?, ?, 0, 1)",
                (email, email, pw_hash),
            )
            user_id = cur.lastrowid
            user_status = "created"
        try:
            db.execute(
                "INSERT INTO session_participants (session_id, user_id, partner, excluded, enrolled_by) "
                "VALUES (?, ?, ?, ?, ?)",
                (session_id, user_id, partner_name, 1 if excluded else 0, enrolled_by_user_id),
            )
            enroll_status = "enrolled"
        except sqlite3.IntegrityError:
            # Already enrolled — reactivate if previously removed
            cur = db.execute(
                "UPDATE session_participants SET removed_at = NULL WHERE session_id = ? AND user_id = ? AND removed_at IS NOT NULL",
                (session_id, user_id),
            )
            enroll_status = "reactivated" if cur.rowcount else "already-enrolled"
            if cur.rowcount:
                any_reactivated = True
        results.append({
            "email": email,
            "partner": partner_name or "Unknown",
            "excluded": excluded,
            "user_status": user_status,
            "enroll_status": enroll_status,
            "initial_password": initial_password,
        })
    db.commit()
    # If any participants were reactivated, their previously-recorded survey
    # responses are now back in the included set — refresh cached counts.
    if any_reactivated:
        for r in db.execute("SELECT id FROM session_surveys WHERE session_id = ?", (session_id,)).fetchall():
            _recompute_survey_counters(r["id"])
    return results


def _session_monitors(session_id):
    return get_db().execute(
        "SELECT sm.id, sm.assigned_at, "
        "       u.id AS user_id, u.username, u.email, u.display_name, u.last_login_at "
        "FROM session_monitors sm JOIN users u ON u.id = sm.user_id "
        "WHERE sm.session_id = ? AND sm.removed_at IS NULL "
        "ORDER BY sm.assigned_at",
        (session_id,),
    ).fetchall()


def _user_is_session_monitor(session_id, user_id):
    if user_id is None:
        return False
    return get_db().execute(
        "SELECT 1 FROM session_monitors WHERE session_id = ? AND user_id = ? AND removed_at IS NULL",
        (session_id, user_id),
    ).fetchone() is not None


def _assign_monitors_to_session(session_id, emails, assigned_by_user_id):
    """For each email: create user if missing, assign as active monitor.
    Mirrors _add_participants_to_session — returns dicts with status + initial_password (or None)."""
    db = get_db()
    results = []
    for email in emails:
        existing = db.execute("SELECT id FROM users WHERE username = ?", (email,)).fetchone()
        if existing:
            user_id = existing["id"]
            initial_password = None
            user_status = "existing"
        else:
            initial_password = secrets.token_urlsafe(10)
            pw_hash = bcrypt.hashpw(initial_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
            cur = db.execute(
                "INSERT INTO users (username, email, password_hash, is_facilitator, is_active) "
                "VALUES (?, ?, ?, 0, 1)",
                (email, email, pw_hash),
            )
            user_id = cur.lastrowid
            user_status = "created"
        try:
            db.execute(
                "INSERT INTO session_monitors (session_id, user_id, assigned_by) "
                "VALUES (?, ?, ?)",
                (session_id, user_id, assigned_by_user_id),
            )
            assign_status = "assigned"
        except sqlite3.IntegrityError:
            cur = db.execute(
                "UPDATE session_monitors SET removed_at = NULL, assigned_at = CURRENT_TIMESTAMP, assigned_by = ? "
                "WHERE session_id = ? AND user_id = ? AND removed_at IS NOT NULL",
                (assigned_by_user_id, session_id, user_id),
            )
            assign_status = "reactivated" if cur.rowcount else "already-assigned"
        results.append({
            "email": email,
            "user_status": user_status,
            "assign_status": assign_status,
            "initial_password": initial_password,
        })
    db.commit()
    return results


def _get_session_or_404(workshop_slug, session_slug):
    row = get_db().execute(
        "SELECT s.*, w.slug AS workshop_slug, w.name AS workshop_name "
        "FROM sessions s JOIN workshops w ON w.id = s.workshop_id "
        "WHERE w.slug = ? AND s.slug = ?",
        (workshop_slug, session_slug),
    ).fetchone()
    if row is None:
        abort(404)
    return row


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/participants", methods=["GET", "POST"])
def admin_session_participants(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    added_results = []
    if request.method == "POST":
        emails = _parse_email_list(request.form.get("emails", ""))
        if emails:
            added_results = _add_participants_to_session(sess["id"], emails, user["id"])
    participants = get_db().execute(
        "SELECT sp.id, sp.partner, sp.excluded, sp.enrolled_at, "
        "       u.id AS user_id, u.username, u.email, u.last_login_at "
        "FROM session_participants sp JOIN users u ON u.id = sp.user_id "
        "WHERE sp.session_id = ? AND sp.removed_at IS NULL "
        "ORDER BY sp.partner IS NULL, sp.partner, u.username",
        (sess["id"],),
    ).fetchall()
    by_partner = {}
    for p in participants:
        key = p["partner"] or "Unknown"
        by_partner.setdefault(key, []).append(p)
    included = sum(1 for p in participants if not p["excluded"])
    reset_creds = session.pop("reset_credentials", None)
    return render_template(
        "admin_session_participants.html",
        user=user, session=sess, by_partner=by_partner,
        total=len(participants), included=included,
        added_results=added_results,
        reset_creds=reset_creds,
    )


# ---------------------------- Survey templates (native) ----------------------------

IFP_BACKGROUND_OPTIONS = [
    "Certified Solution Architect",
    "Certified Model Builder",
    "Certified Master Anaplanner",
    "Program Architect",
    "Practice Lead",
    "Project Manager",
]
IFP_EXPERIENCE_OPTIONS = ["No Projects (yet)", "2 to 5 Projects", "More than 5 Projects"]
IFP_TOPICS = [
    "Anaplan Way Methodology",
    "Application Lifecycle Management (ALM)",
    "Anaplan Application Configuration",
    "Integrated Financial Planning (IFP) Application",
    "Extensions to IFP",
    "Anaplan Data Orchestrator (ADO)",
]

IFP_PRE_TEMPLATE = {
    "intro": "Take a few minutes before the workshop starts to share your background and where you're starting from. Your answers help us calibrate the live discussions and the wrap-up retrospective.",
    "questions": (
        [
            {"id": "background_role", "text": "Your background and role (select all that apply)", "type": "multi_select", "options": IFP_BACKGROUND_OPTIONS, "required": False},
            {"id": "project_experience", "text": "Your project experience", "type": "single_select", "options": IFP_EXPERIENCE_OPTIONS, "required": True},
        ]
        + [{"id": f"topic_{i}", "text": t, "type": "likert", "required": True} for i, t in enumerate(IFP_TOPICS)]
        + [{"id": "learning_hopes", "text": "Add one or two things you hope to learn from the course", "type": "free_text", "required": False}]
    ),
}

IFP_POST_TEMPLATE = {
    "intro": "Now that the workshop is wrapping up, share how things changed and what we should improve for the next cohort.",
    "questions": (
        [{"id": f"topic_{i}", "text": t, "type": "likert", "required": True} for i, t in enumerate(IFP_TOPICS)]
        + [
            {"id": "feedback_valuable", "text": "What did you find most valuable about the workshop?", "type": "free_text", "required": False},
            {"id": "feedback_improve", "text": "What aspects of the workshop could be improved?", "type": "free_text", "required": False},
            {"id": "feedback_topics", "text": "Are there additional topics you would like covered in future sessions?", "type": "free_text", "required": False},
            {"id": "feedback_instructors", "text": "What feedback do you have for the instructors?", "type": "free_text", "required": False},
        ]
    ),
}


def _load_template(workshop_id, kind):
    row = get_db().execute(
        "SELECT * FROM survey_templates WHERE workshop_id = ? AND kind = ?",
        (workshop_id, kind),
    ).fetchone()
    if not row:
        return None
    keys = set(row.keys())
    return {
        "id": row["id"],
        "workshop_id": row["workshop_id"],
        "kind": row["kind"],
        "intro": row["intro"],
        "questions": json.loads(row["questions_json"]),
        "passing_score": row["passing_score"] if "passing_score" in keys else 70,
        "max_attempts": row["max_attempts"] if "max_attempts" in keys else None,
        "questions_per_attempt": row["questions_per_attempt"] if "questions_per_attempt" in keys else None,
        "shuffle_options": bool(row["shuffle_options"]) if "shuffle_options" in keys else True,
        "updated_at": row["updated_at"],
    }


def _save_template(workshop_id, kind, intro, questions, updated_by,
                   passing_score=70, max_attempts=None,
                   questions_per_attempt=None, shuffle_options=True):
    db = get_db()
    db.execute(
        "INSERT INTO survey_templates (workshop_id, kind, intro, questions_json, "
        "passing_score, max_attempts, questions_per_attempt, shuffle_options, updated_by) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?) "
        "ON CONFLICT(workshop_id, kind) DO UPDATE SET "
        "intro = excluded.intro, questions_json = excluded.questions_json, "
        "passing_score = excluded.passing_score, max_attempts = excluded.max_attempts, "
        "questions_per_attempt = excluded.questions_per_attempt, "
        "shuffle_options = excluded.shuffle_options, "
        "updated_at = CURRENT_TIMESTAMP, updated_by = excluded.updated_by",
        (workshop_id, kind, intro, json.dumps(questions),
         passing_score, max_attempts, questions_per_attempt, 1 if shuffle_options else 0, updated_by),
    )
    db.commit()


def _draw_kc_attempt(template):
    """Random-select N questions from the pool and optionally shuffle their options.
    Returns a new template dict with `questions` replaced by the drawn subset.
    Each question is deep-copied so the pool itself isn't mutated. Correct value
    is option text, so option shuffle doesn't affect scoring downstream."""
    import random
    pool = template.get("questions", [])
    n = template.get("questions_per_attempt")
    if n is not None and 0 < n < len(pool):
        refs = random.sample(pool, n)
    else:
        refs = list(pool)
        random.shuffle(refs)
    chosen = [dict(q) for q in refs]  # shallow copy each q so option shuffle stays local
    if template.get("shuffle_options", True):
        for q in chosen:
            if q.get("type") in ("single_select", "multi_select") and isinstance(q.get("options"), list):
                opts = list(q["options"])
                random.shuffle(opts)
                q["options"] = opts
    return {**template, "questions": chosen}


def _kc_user_status(session_id, user_id):
    """Return (attempts_so_far, has_passed) for this user's knowledge-check submissions in this session."""
    row = get_db().execute(
        "SELECT COUNT(*) AS n, MAX(COALESCE(sr.passed, 0)) AS any_passed "
        "FROM survey_responses sr JOIN session_surveys ss ON ss.id = sr.survey_id "
        "WHERE ss.session_id = ? AND ss.kind = 'knowledge' AND sr.participant_user_id = ?",
        (session_id, user_id),
    ).fetchone()
    return (row["n"] or 0, bool(row["any_passed"] or 0))


def _score_knowledge_response(template, responses_by_text):
    """Score a knowledge-check submission against the template's correct answers.
    Returns (score_pct, passed_bool, total_scoreable). Only single_select and
    multi_select questions with a `correct` field are scoreable."""
    correct_count = 0
    total = 0
    for q in template["questions"]:
        if q.get("type") not in ("single_select", "multi_select"):
            continue
        correct = q.get("correct")
        if correct is None or correct == "" or correct == []:
            continue
        total += 1
        ans = responses_by_text.get(q["text"], "")
        if q["type"] == "single_select":
            if isinstance(correct, list):
                correct = correct[0] if correct else ""
            if str(ans).strip().lower() == str(correct).strip().lower():
                correct_count += 1
        else:  # multi_select — answer is "A;B;" semicolon-delimited
            ans_set = {p.strip().lower() for p in str(ans).split(";") if p.strip()}
            if isinstance(correct, list):
                correct_set = {str(c).strip().lower() for c in correct if str(c).strip()}
            else:
                correct_set = {p.strip().lower() for p in str(correct).split(";") if p.strip()}
            if ans_set == correct_set:
                correct_count += 1
    if total == 0:
        return None, None, 0
    score_pct = round(correct_count / total * 100)
    passed = score_pct >= int(template.get("passing_score", 70))
    return score_pct, passed, total


def _ensure_session_survey(session_id, kind):
    """Return the session_surveys row id for (session_id, kind), creating a thin shell row if absent."""
    db = get_db()
    row = db.execute(
        "SELECT id FROM session_surveys WHERE session_id = ? AND kind = ?",
        (session_id, kind),
    ).fetchone()
    if row:
        return row["id"]
    cur = db.execute(
        "INSERT INTO session_surveys (session_id, kind, filename, total_rows, is_anonymous, headers_json) "
        "VALUES (?, ?, NULL, 0, 0, '[]')",
        (session_id, kind),
    )
    db.commit()
    return cur.lastrowid


def _recompute_survey_counters(survey_id):
    """Update cached counters on session_surveys, applying the same enrollment
    filter that _load_survey_responses uses so the counts stay consistent."""
    db = get_db()
    rows = db.execute(
        "SELECT COUNT(*) AS n, COUNT(sr.participant_user_id) AS named "
        "FROM survey_responses sr "
        "JOIN session_surveys ss ON ss.id = sr.survey_id "
        "WHERE sr.survey_id = ? "
        "  AND ( "
        "    sr.participant_user_id IS NULL "
        "    OR EXISTS ( "
        "      SELECT 1 FROM session_participants sp "
        "      WHERE sp.session_id = ss.session_id "
        "        AND sp.user_id = sr.participant_user_id "
        "        AND sp.removed_at IS NULL "
        "    ) "
        "  )",
        (survey_id,),
    ).fetchone()
    is_anon = 1 if (rows["n"] > 0 and rows["named"] == 0) else 0
    db.execute(
        "UPDATE session_surveys SET total_rows = ?, is_anonymous = ? WHERE id = ?",
        (rows["n"], is_anon, survey_id),
    )
    db.commit()


# ---------------------------- Surveys ----------------------------

def _coerce_cell(v):
    """Coerce a cell value into something JSON-serializable. Mostly: datetime → ISO string."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat(sep=" ", timespec="seconds")
    if isinstance(v, date):
        return v.isoformat()
    return v


def _parse_xlsx_upload(file_storage):
    """Parse uploaded XLSX into (headers, rows). Rows is list[dict header->value].
    Uses non-read-only mode so files with stale max_row/max_column metadata
    (a Microsoft Forms export quirk) are still parsed correctly."""
    from openpyxl import load_workbook
    wb = load_workbook(file_storage, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    headers = []
    seen_h = set()
    for i, h in enumerate(all_rows[0] or []):
        name = str(h).strip() if h is not None else f"col_{i}"
        if name in seen_h:
            name = f"{name}_{i}"
        seen_h.add(name)
        headers.append(name)
    rows = []
    for row in all_rows[1:]:
        if not row or not any(c not in (None, "") for c in row):
            continue
        rows.append({headers[i]: _coerce_cell(row[i]) for i in range(min(len(headers), len(row)))})
    return headers, rows


def _parse_csv_upload(file_storage):
    """Parse uploaded CSV into (headers, rows). Same shape as _parse_xlsx_upload."""
    import csv
    import io
    raw = file_storage.read()
    if isinstance(raw, bytes):
        # Strip BOM if present, decode utf-8
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="replace")
    else:
        text = raw
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        return [], []
    headers = []
    seen_h = set()
    for i, h in enumerate(all_rows[0]):
        name = (h or "").strip() or f"col_{i}"
        if name in seen_h:
            name = f"{name}_{i}"
        seen_h.add(name)
        headers.append(name)
    rows = []
    for row in all_rows[1:]:
        if not row or not any((c or "").strip() for c in row):
            continue
        rows.append({headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))})
    return headers, rows


def _is_metadata_header(h):
    return str(h).strip().lower() in SURVEY_METADATA_HEADERS


def _detect_question_type(values):
    """Given non-empty values, return 'likert' | 'multi_select' | 'single_select' | 'pass_fail' | 'free_text'.
    Multi-select and single-select both require the underlying options to look curated —
    a small finite set of short labels. Verbose prose answers (even with semicolons or
    short repeating phrasing) fall through to free_text."""
    cleaned = [str(v).strip() for v in values if v is not None and str(v).strip()]
    if not cleaned:
        return "free_text"
    if any(";" in v for v in cleaned):
        # Split on ; and check the underlying option set is finite and short.
        parts = []
        for v in cleaned:
            parts.extend(p.strip() for p in v.split(";") if p.strip())
        if parts:
            unique_parts = {p.lower() for p in parts}
            max_part_len = max(len(p) for p in parts)
            if len(unique_parts) <= 8 and max_part_len <= 50:
                return "multi_select"
    lower = [v.lower() for v in cleaned]
    if all(v in LIKERT_MAP for v in lower):
        return "likert"
    if all(PASS_FAIL_RE.match(v) or FAIL_RE.match(v) for v in cleaned):
        return "pass_fail"
    unique = set(lower)
    diversity = len(unique) / len(cleaned)
    max_len = max(len(v) for v in cleaned)
    if len(unique) <= 6 and max_len <= 50 and diversity <= 0.4:
        return "single_select"
    return "free_text"


def _normalize_categorical(v):
    s = str(v).strip()
    return CATEGORICAL_ALIASES.get(s.lower(), s)


def _split_multi(v):
    return [_normalize_categorical(p) for p in str(v).split(";") if p.strip()]


def _detect_anonymous(rows):
    has_email = any(str(r.get("Email", "") or "").strip().lower() not in ("", "anonymous") and "@" in str(r.get("Email", "") or "") for r in rows)
    return not has_email and len(rows) > 0


def _aggregate_question(qheader, qtype, rows):
    """Compute aggregate stats for one question across rows."""
    values = []
    for r in rows:
        v = r["responses"].get(qheader)
        if v is None or str(v).strip() == "":
            continue
        values.append(v)
    if qtype == "likert":
        nums = []
        dist = {label: 0 for label in LIKERT_ORDER}
        for v in values:
            label_key = str(v).strip().lower()
            n = LIKERT_MAP.get(label_key)
            if n is None:
                continue
            nums.append(n)
            display_label = str(v).strip()
            if display_label not in dist:
                dist[display_label] = 0
            dist[display_label] += 1
        avg = round(sum(nums) / len(nums), 2) if nums else None
        ordered = [{"label": k, "count": v, "pct": round(v / len(nums) * 100, 1) if nums else 0} for k, v in dist.items()]
        return {"type": "likert", "n": len(nums), "avg": avg, "distribution": ordered}
    if qtype == "multi_select":
        counts = {}
        for v in values:
            for opt in _split_multi(v):
                counts[opt] = counts.get(opt, 0) + 1
        n_respondents = len(values)
        items = sorted(counts.items(), key=lambda x: -x[1])
        return {
            "type": "multi_select",
            "n": n_respondents,
            "items": [{"label": k, "count": v, "pct": round(v / n_respondents * 100, 1) if n_respondents else 0} for k, v in items],
        }
    if qtype == "single_select":
        counts = {}
        for v in values:
            opt = _normalize_categorical(v)
            counts[opt] = counts.get(opt, 0) + 1
        items = sorted(counts.items(), key=lambda x: -x[1])
        return {
            "type": "single_select",
            "n": len(values),
            "items": [{"label": k, "count": v, "pct": round(v / len(values) * 100, 1)} for k, v in items],
        }
    if qtype == "pass_fail":
        passed = sum(1 for v in values if PASS_FAIL_RE.match(str(v).strip()))
        failed = len(values) - passed
        return {"type": "pass_fail", "n": len(values), "passed": passed, "failed": failed, "pass_rate": round(passed / len(values) * 100, 1) if values else 0}
    samples = [str(v) for v in values[:200]]
    return {"type": "free_text", "n": len(values), "samples": samples}


def _store_survey_upload(session_id, kind, filename, uploaded_by, headers, rows):
    """Persist the parsed survey. Replaces any existing same-kind survey for this session."""
    db = get_db()
    is_anon = _detect_anonymous(rows)
    db.execute("DELETE FROM session_surveys WHERE session_id = ? AND kind = ?", (session_id, kind))
    cur = db.execute(
        "INSERT INTO session_surveys (session_id, kind, filename, uploaded_by, total_rows, is_anonymous, headers_json) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (session_id, kind, filename, uploaded_by, len(rows), 1 if is_anon else 0, json.dumps(headers)),
    )
    survey_id = cur.lastrowid
    for r in rows:
        email = str(r.get("Email", "") or "").strip().lower()
        if not email or email == "anonymous" or "@" not in email:
            email_to_store = None
            user_id = None
        else:
            email_to_store = email
            u = db.execute("SELECT id FROM users WHERE username = ?", (email,)).fetchone()
            user_id = u["id"] if u else None
        db.execute(
            "INSERT INTO survey_responses (survey_id, participant_user_id, raw_email, responses_json) "
            "VALUES (?, ?, ?, ?)",
            (survey_id, user_id, email_to_store, json.dumps(r)),
        )
    db.commit()
    return survey_id, is_anon


def _load_session_surveys(session_id):
    """Return dict {kind: survey_row} for this session."""
    rows = get_db().execute(
        "SELECT * FROM session_surveys WHERE session_id = ?", (session_id,)
    ).fetchall()
    return {r["kind"]: dict(r) for r in rows}


def _load_survey_responses(survey_id):
    """Load responses for a survey, EXCLUDING responses from users whose enrollment
    in the survey's session has been soft-removed. Anonymous responses (no user_id)
    are always included since they have no enrollment to check."""
    rows = get_db().execute(
        "SELECT sr.participant_user_id, sr.raw_email, sr.responses_json, sr.score, sr.passed "
        "FROM survey_responses sr "
        "JOIN session_surveys ss ON ss.id = sr.survey_id "
        "WHERE sr.survey_id = ? "
        "  AND ( "
        "    sr.participant_user_id IS NULL "
        "    OR EXISTS ( "
        "      SELECT 1 FROM session_participants sp "
        "      WHERE sp.session_id = ss.session_id "
        "        AND sp.user_id = sr.participant_user_id "
        "        AND sp.removed_at IS NULL "
        "    ) "
        "  )",
        (survey_id,),
    ).fetchall()
    out = []
    for r in rows:
        out.append({
            "user_id": r["participant_user_id"],
            "raw_email": r["raw_email"],
            "responses": json.loads(r["responses_json"]),
            "score": r["score"],
            "passed": r["passed"],
        })
    return out


def _classify_columns(headers, responses):
    """Return list of {header, qtype} for non-metadata columns."""
    out = []
    for h in headers:
        if _is_metadata_header(h):
            continue
        vals = [r["responses"].get(h) for r in responses]
        out.append({"header": h, "qtype": _detect_question_type(vals)})
    return out


@app.route("/admin/workshops/<slug>/survey-templates")
def admin_survey_templates(slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    workshop = get_db().execute("SELECT * FROM workshops WHERE slug = ?", (slug,)).fetchone()
    if workshop is None:
        abort(404)
    templates = {
        "pre": _load_template(workshop["id"], "pre"),
        "post": _load_template(workshop["id"], "post"),
        "knowledge": _load_template(workshop["id"], "knowledge"),
    }
    return render_template(
        "admin_survey_templates.html",
        user=user, workshop=workshop, templates=templates,
    )


@app.route("/admin/workshops/<slug>/survey-templates/<kind>/import", methods=["POST"])
def admin_survey_template_import(slug, kind):
    """Append a JSON-array of questions to the existing pool."""
    user, redir = _require_facilitator()
    if redir:
        return redir
    if kind not in ("pre", "post", "knowledge"):
        abort(404)
    workshop = get_db().execute("SELECT * FROM workshops WHERE slug = ?", (slug,)).fetchone()
    if workshop is None:
        abort(404)
    template = _load_template(workshop["id"], kind)
    raw = request.form.get("questions_json", "")
    try:
        incoming = json.loads(raw)
    except json.JSONDecodeError:
        return redirect(url_for("admin_survey_template_edit", slug=slug, kind=kind))
    if not isinstance(incoming, list):
        return redirect(url_for("admin_survey_template_edit", slug=slug, kind=kind))
    existing = list(template["questions"]) if template else []
    existing_ids = {q.get("id") for q in existing}
    next_idx = len(existing) + 1
    for q in incoming:
        if not isinstance(q, dict):
            continue
        text = str(q.get("text", "")).strip()
        qtype = q.get("type")
        if not text or qtype not in ("likert", "multi_select", "single_select", "free_text"):
            continue
        qid = str(q.get("id") or "").strip()
        if not qid or qid in existing_ids:
            qid = f"q{next_idx}"
            while qid in existing_ids:
                next_idx += 1
                qid = f"q{next_idx}"
        existing_ids.add(qid)
        next_idx += 1
        entry = {"id": qid, "text": text, "type": qtype, "required": bool(q.get("required", False))}
        if qtype in ("multi_select", "single_select"):
            opts = q.get("options") or []
            entry["options"] = [str(o).strip() for o in opts if str(o).strip()]
            if kind == "knowledge":
                correct_raw = q.get("correct")
                if qtype == "single_select":
                    cstr = str(correct_raw or "").strip()
                    if cstr:
                        entry["correct"] = cstr
                else:
                    if isinstance(correct_raw, list):
                        clist = [str(c).strip() for c in correct_raw if str(c).strip()]
                    else:
                        clist = [p.strip() for p in str(correct_raw or "").split(";") if p.strip()]
                    if clist:
                        entry["correct"] = clist
        existing.append(entry)
    _save_template(
        workshop["id"], kind,
        template["intro"] if template else "",
        existing, user["id"],
        passing_score=template["passing_score"] if template else 70,
        max_attempts=template["max_attempts"] if template else None,
        questions_per_attempt=template["questions_per_attempt"] if template else None,
        shuffle_options=template["shuffle_options"] if template else True,
    )
    return redirect(url_for("admin_survey_template_edit", slug=slug, kind=kind))


@app.route("/admin/workshops/<slug>/survey-templates/<kind>/edit", methods=["GET", "POST"])
def admin_survey_template_edit(slug, kind):
    user, redir = _require_facilitator()
    if redir:
        return redir
    if kind not in ("pre", "post", "knowledge"):
        abort(404)
    workshop = get_db().execute("SELECT * FROM workshops WHERE slug = ?", (slug,)).fetchone()
    if workshop is None:
        abort(404)

    error = None
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        if action == "load_ifp":
            seed = IFP_PRE_TEMPLATE if kind == "pre" else (IFP_POST_TEMPLATE if kind == "post" else {"intro": "", "questions": []})
            _save_template(workshop["id"], kind, seed["intro"], seed["questions"], user["id"])
            return redirect(url_for("admin_survey_template_edit", slug=slug, kind=kind))
        if action == "delete":
            get_db().execute("DELETE FROM survey_templates WHERE workshop_id = ? AND kind = ?", (workshop["id"], kind))
            get_db().commit()
            return redirect(url_for("admin_survey_templates", slug=slug))
        intro = (request.form.get("intro") or "").strip()
        try:
            passing_score = int(request.form.get("passing_score") or 70)
        except ValueError:
            passing_score = 70
        passing_score = max(0, min(100, passing_score))
        max_attempts_raw = (request.form.get("max_attempts") or "").strip()
        if max_attempts_raw == "":
            max_attempts = None
        else:
            try:
                max_attempts = max(1, int(max_attempts_raw))
            except ValueError:
                max_attempts = None
        qpa_raw = (request.form.get("questions_per_attempt") or "").strip()
        if qpa_raw == "":
            questions_per_attempt = None
        else:
            try:
                questions_per_attempt = max(1, int(qpa_raw))
            except ValueError:
                questions_per_attempt = None
        shuffle_options = request.form.get("shuffle_options") == "on"
        try:
            questions = json.loads(request.form.get("questions_json") or "[]")
        except json.JSONDecodeError:
            error = "Invalid questions data."
            questions = []
        if isinstance(questions, list):
            cleaned = []
            for i, q in enumerate(questions):
                if not isinstance(q, dict):
                    continue
                text = str(q.get("text", "")).strip()
                qtype = q.get("type")
                if not text or qtype not in ("likert", "multi_select", "single_select", "free_text"):
                    continue
                entry = {"id": q.get("id") or f"q{i+1}", "text": text, "type": qtype, "required": bool(q.get("required", False))}
                if qtype in ("multi_select", "single_select"):
                    opts = q.get("options") or []
                    entry["options"] = [str(o).strip() for o in opts if str(o).strip()]
                    if kind == "knowledge":
                        correct_raw = q.get("correct")
                        if qtype == "single_select":
                            cstr = str(correct_raw or "").strip()
                            if cstr:
                                entry["correct"] = cstr
                        else:
                            if isinstance(correct_raw, list):
                                clist = [str(c).strip() for c in correct_raw if str(c).strip()]
                            else:
                                clist = [p.strip() for p in str(correct_raw or "").split(";") if p.strip()]
                            if clist:
                                entry["correct"] = clist
                cleaned.append(entry)
            questions = cleaned
        if not error and not questions:
            error = "At least one question is required."
        if not error:
            _save_template(workshop["id"], kind, intro, questions, user["id"],
                           passing_score=passing_score, max_attempts=max_attempts,
                           questions_per_attempt=questions_per_attempt,
                           shuffle_options=shuffle_options if kind == "knowledge" else True)
            return redirect(url_for("admin_survey_templates", slug=slug))

    template = _load_template(workshop["id"], kind)
    return render_template(
        "admin_survey_template_edit.html",
        user=user, workshop=workshop, kind=kind, template=template, error=error,
        has_ifp_seed=(kind in ("pre", "post")),
    )


# ---------------------------- Participant survey form ----------------------------

def _current_enrollment_session(user_id, workshop_slug):
    """Most recent (non-removed, non-archived) session enrollment for this user in this workshop."""
    return get_db().execute(
        "SELECT s.id, s.slug, s.name FROM session_participants sp "
        "JOIN sessions s ON s.id = sp.session_id "
        "JOIN workshops w ON w.id = s.workshop_id "
        "WHERE sp.user_id = ? AND sp.removed_at IS NULL "
        "  AND w.slug = ? AND s.status != 'archived' "
        "ORDER BY sp.enrolled_at DESC LIMIT 1",
        (user_id, workshop_slug),
    ).fetchone()


@app.route("/w/<workshop_slug>/survey/<kind>", methods=["GET", "POST"])
def take_survey(workshop_slug, kind):
    if kind not in ("pre", "post", "knowledge"):
        abort(404)
    user = current_user()
    if user is None:
        return redirect(url_for("login", next=request.path))
    workshop = get_db().execute("SELECT * FROM workshops WHERE slug = ?", (workshop_slug,)).fetchone()
    if workshop is None:
        abort(404)
    template = _load_template(workshop["id"], kind)
    if template is None:
        return render_template("survey_unavailable.html", workshop=workshop, kind=kind, reason="Your facilitator hasn't published this survey yet."), 404
    enrollment = _current_enrollment_session(user["id"], workshop_slug)
    if enrollment is None and not user["is_facilitator"]:
        return render_template("survey_unavailable.html", workshop=workshop, kind=kind, reason="We couldn't find an active session enrollment for you in this workshop. Reach out to your facilitator."), 403

    # Gating + retry logic for knowledge-check kind (allows retakes until passed or limit reached).
    attempts_so_far = 0
    has_passed = False
    attempts_left = None
    if kind == "knowledge" and enrollment is not None:
        attempts_so_far, has_passed = _kc_user_status(enrollment["id"], user["id"])
        max_attempts = template.get("max_attempts")
        if max_attempts is not None:
            attempts_left = max(0, max_attempts - attempts_so_far)
        if has_passed:
            return render_template(
                "survey_already_passed.html",
                workshop=workshop, kind=kind, session_name=enrollment["name"],
                attempts_so_far=attempts_so_far, max_attempts=max_attempts,
            )
        if max_attempts is not None and attempts_so_far >= max_attempts:
            return render_template(
                "survey_no_more_attempts.html",
                workshop=workshop, kind=kind, session_name=enrollment["name"],
                attempts_so_far=attempts_so_far, max_attempts=max_attempts,
            )

    # For knowledge checks, GET draws a fresh random subset from the pool.
    # POST honors whichever subset was submitted (read from a hidden q_ids field).
    rendered_template = template
    if kind == "knowledge":
        if request.method == "POST":
            presented_ids = [pid for pid in (request.form.get("q_ids") or "").split(",") if pid]
            if presented_ids:
                drawn = [q for q in template["questions"] if q.get("id") in presented_ids]
                rendered_template = {**template, "questions": drawn}
            else:
                rendered_template = _draw_kc_attempt(template)
        else:
            rendered_template = _draw_kc_attempt(template)
    template = rendered_template

    error = None
    if request.method == "POST":
        if enrollment is None:
            abort(403)
        if kind != "knowledge":
            already = get_db().execute(
                "SELECT sr.id FROM survey_responses sr "
                "JOIN session_surveys ss ON ss.id = sr.survey_id "
                "WHERE ss.session_id = ? AND ss.kind = ? AND sr.participant_user_id = ?",
                (enrollment["id"], kind, user["id"]),
            ).fetchone()
            if already:
                return render_template("survey_already_submitted.html", workshop=workshop, kind=kind, session_name=enrollment["name"])
        responses = {}
        missing = []
        for q in template["questions"]:
            qtype = q["type"]
            if qtype == "multi_select":
                vals = request.form.getlist(f"answer__{q['id']}")
                responses[q["text"]] = ";".join(vals) + ";" if vals else ""
                if q.get("required") and not vals:
                    missing.append(q["text"])
            else:
                v = (request.form.get(f"answer__{q['id']}") or "").strip()
                responses[q["text"]] = v
                if q.get("required") and not v:
                    missing.append(q["text"])
        if missing:
            error = "Please answer all required questions: " + "; ".join(missing[:5]) + (" ..." if len(missing) > 5 else "")
        else:
            survey_id = _ensure_session_survey(enrollment["id"], kind)
            score, passed, scoreable = (None, None, 0)
            if kind == "knowledge":
                score, passed, scoreable = _score_knowledge_response(template, responses)
            this_attempt = (attempts_so_far + 1) if kind == "knowledge" else None
            get_db().execute(
                "INSERT INTO survey_responses (survey_id, participant_user_id, raw_email, responses_json, score, passed, attempt_number, submitted_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)",
                (survey_id, user["id"], user["username"], json.dumps(responses),
                 score, (1 if passed else (0 if passed is False else None)), this_attempt),
            )
            get_db().commit()
            _recompute_survey_counters(survey_id)
            max_attempts = template.get("max_attempts") if kind == "knowledge" else None
            attempts_remaining = None
            if max_attempts is not None and this_attempt is not None:
                attempts_remaining = max(0, max_attempts - this_attempt)
            return render_template(
                "survey_thanks.html",
                workshop=workshop, kind=kind, session_name=enrollment["name"],
                score=score, passed=passed, scoreable=scoreable,
                passing_score=template.get("passing_score", 70),
                this_attempt=this_attempt, max_attempts=max_attempts,
                attempts_remaining=attempts_remaining,
            )

    return render_template(
        "survey_form.html",
        workshop=workshop, kind=kind, template=template,
        likert_order=LIKERT_ORDER,
        session_name=enrollment["name"] if enrollment else None,
        is_facilitator_preview=(user["is_facilitator"] and enrollment is None),
        error=error,
        attempts_so_far=attempts_so_far,
        attempts_left=attempts_left,
    )


# ---------------------------- Session reports ----------------------------

def _partner_slug(name):
    s = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
    return s or "unknown"


def _matched_likert_table(pre_responses, post_responses):
    """Pre/post Likert table matched by question text. Reused from session surveys page."""
    if not pre_responses or not post_responses:
        return []
    pre_headers = set()
    for r in pre_responses:
        pre_headers.update(r["responses"].keys())
    cols = []
    for h in pre_headers:
        if _is_metadata_header(h):
            continue
        vals = [r["responses"].get(h) for r in pre_responses]
        if _detect_question_type([v for v in vals if v]) != "likert":
            continue
        cols.append(h)
    out = []
    for h in cols:
        pre_agg = _aggregate_question(h, "likert", pre_responses)
        post_agg = _aggregate_question(h, "likert", post_responses)
        if pre_agg["avg"] is None or post_agg["avg"] is None:
            continue
        out.append({
            "header": h,
            "pre_avg": pre_agg["avg"], "post_avg": post_agg["avg"],
            "delta": round(post_agg["avg"] - pre_agg["avg"], 2),
            "pre_n": pre_agg["n"], "post_n": post_agg["n"],
        })
    out.sort(key=lambda x: -x["delta"])
    return out


def _session_report_data(session_id):
    """Compute every metric the reports need from one query batch."""
    db = get_db()
    session = db.execute(
        "SELECT s.*, w.slug AS workshop_slug, w.name AS workshop_name "
        "FROM sessions s JOIN workshops w ON w.id = s.workshop_id "
        "WHERE s.id = ?", (session_id,),
    ).fetchone()
    participants = db.execute(
        "SELECT sp.user_id, sp.partner, sp.excluded, u.username, u.email "
        "FROM session_participants sp JOIN users u ON u.id = sp.user_id "
        "WHERE sp.session_id = ? AND sp.removed_at IS NULL "
        "ORDER BY sp.partner IS NULL, sp.partner, u.username",
        (session_id,),
    ).fetchall()
    surveys = _load_session_surveys(session_id)
    pre_resp = _load_survey_responses(surveys["pre"]["id"]) if "pre" in surveys else []
    post_resp = _load_survey_responses(surveys["post"]["id"]) if "post" in surveys else []
    kc_resp = _load_survey_responses(surveys["knowledge"]["id"]) if "knowledge" in surveys else []

    pre_users = {r["user_id"] for r in pre_resp if r["user_id"]}
    post_users = {r["user_id"] for r in post_resp if r["user_id"]}
    kc_pass_users = set()
    for r in kc_resp:
        if not r["user_id"]:
            continue
        # Prefer the stored `passed` column (native scored knowledge checks).
        if r.get("passed") == 1:
            kc_pass_users.add(r["user_id"])
            continue
        if r.get("passed") == 0:
            continue
        # Fallback for legacy uploads with raw Pass/Fail strings in the responses.
        for v in r["responses"].values():
            if v and PASS_FAIL_RE.match(str(v).strip()):
                kc_pass_users.add(r["user_id"])
                break

    counted = [p for p in participants if not p["excluded"]]
    excluded = [p for p in participants if p["excluded"]]
    partners = {}
    for p in counted:
        name = p["partner"] or "Unknown"
        bucket = partners.setdefault(name, {
            "name": name, "slug": _partner_slug(name),
            "people": [], "pre_done": 0, "post_done": 0, "kc_pass": 0, "all_three": 0,
        })
        bucket["people"].append(p)
        if p["user_id"] in pre_users: bucket["pre_done"] += 1
        if p["user_id"] in post_users: bucket["post_done"] += 1
        if p["user_id"] in kc_pass_users: bucket["kc_pass"] += 1
        if (p["user_id"] in pre_users and p["user_id"] in post_users and p["user_id"] in kc_pass_users):
            bucket["all_three"] += 1
    for b in partners.values():
        b["count"] = len(b["people"])
        b["rate"] = (b["all_three"] / b["count"]) if b["count"] else 0
    ranked = sorted(partners.values(), key=lambda b: (-b["rate"], -b["count"], b["name"]))

    standouts = [b for b in ranked if b["count"] >= 1 and b["rate"] == 1.0]
    at_risk = [b for b in ranked if b["count"] >= 1 and b["rate"] == 0.0
               and (b["pre_done"] == 0 and b["post_done"] == 0 and b["kc_pass"] == 0)]

    matched_likert = _matched_likert_table(pre_resp, post_resp)
    biggest_gain = matched_likert[0] if matched_likert and matched_likert[0]["delta"] > 0 else None
    pre_likert_avg, _ = _likert_avg_across_responses(pre_resp)
    post_likert_avg, _ = _likert_avg_across_responses(post_resp)
    overall_delta = round(post_likert_avg - pre_likert_avg, 2) if (pre_likert_avg is not None and post_likert_avg is not None) else None

    pre_count = sum(1 for p in counted if p["user_id"] in pre_users)
    post_count = sum(1 for p in counted if p["user_id"] in post_users)
    kc_pass_count = sum(1 for p in counted if p["user_id"] in kc_pass_users)
    all_three = sum(1 for p in counted
                    if p["user_id"] in pre_users and p["user_id"] in post_users and p["user_id"] in kc_pass_users)
    total_counted = len(counted)

    return {
        "session": dict(session) if session else None,
        "total_registered": len(participants),
        "total_counted": total_counted,
        "total_excluded": len(excluded),
        "pre_count": pre_count, "post_count": post_count,
        "kc_pass_count": kc_pass_count, "all_three": all_three,
        "pre_rate": round(pre_count / total_counted * 100) if total_counted else 0,
        "post_rate": round(post_count / total_counted * 100) if total_counted else 0,
        "kc_rate": round(kc_pass_count / total_counted * 100) if total_counted else 0,
        "all_three_rate": round(all_three / total_counted * 100) if total_counted else 0,
        "partners": ranked,
        "partner_count": len(partners),
        "standouts": standouts, "at_risk": at_risk,
        "pre_likert_avg": pre_likert_avg, "post_likert_avg": post_likert_avg,
        "overall_delta": overall_delta,
        "matched_likert": matched_likert, "biggest_gain": biggest_gain,
        "today": date.today().isoformat(),
    }


def _workshop_rollup_data(workshop_id, workshop_name, workshop_slug):
    """Aggregate report data across every non-archived session of a workshop."""
    db = get_db()
    sessions = db.execute(
        "SELECT id, slug, name, start_date, end_date "
        "FROM sessions WHERE workshop_id = ? AND status != 'archived' "
        "ORDER BY start_date",
        (workshop_id,),
    ).fetchall()
    if not sessions:
        return None
    session_data = [_session_report_data(s["id"]) for s in sessions]

    total_registered = sum(s["total_registered"] for s in session_data)
    total_counted = sum(s["total_counted"] for s in session_data)
    pre_count = sum(s["pre_count"] for s in session_data)
    post_count = sum(s["post_count"] for s in session_data)
    kc_pass_count = sum(s["kc_pass_count"] for s in session_data)
    all_three = sum(s["all_three"] for s in session_data)

    from collections import defaultdict
    partner_totals = defaultdict(lambda: {"count": 0, "pre_done": 0, "post_done": 0, "kc_pass": 0, "all_three": 0})
    for sd in session_data:
        for b in sd["partners"]:
            agg = partner_totals[b["name"]]
            agg["count"] += b["count"]
            agg["pre_done"] += b["pre_done"]
            agg["post_done"] += b["post_done"]
            agg["kc_pass"] += b["kc_pass"]
            agg["all_three"] += b["all_three"]
    partners = []
    for name, agg in partner_totals.items():
        rate = (agg["all_three"] / agg["count"]) if agg["count"] else 0
        partners.append({"name": name, "slug": _partner_slug(name),
                         "count": agg["count"], "all_three": agg["all_three"],
                         "pre_done": agg["pre_done"], "post_done": agg["post_done"],
                         "kc_pass": agg["kc_pass"], "rate": rate})
    partners.sort(key=lambda b: (-b["rate"], -b["count"], b["name"]))

    pre_avgs = [s["pre_likert_avg"] for s in session_data if s["pre_likert_avg"] is not None]
    post_avgs = [s["post_likert_avg"] for s in session_data if s["post_likert_avg"] is not None]
    workshop_pre_avg = round(sum(pre_avgs) / len(pre_avgs), 2) if pre_avgs else None
    workshop_post_avg = round(sum(post_avgs) / len(post_avgs), 2) if post_avgs else None
    workshop_delta = (round(workshop_post_avg - workshop_pre_avg, 2)
                      if workshop_pre_avg is not None and workshop_post_avg is not None else None)

    return {
        "workshop_name": workshop_name,
        "workshop_slug": workshop_slug,
        "session_count": len(sessions),
        "first_date": sessions[0]["start_date"],
        "last_date": sessions[-1]["end_date"],
        "total_registered": total_registered,
        "total_counted": total_counted,
        "partner_count": len(partner_totals),
        "pre_count": pre_count, "post_count": post_count,
        "kc_pass_count": kc_pass_count, "all_three": all_three,
        "pre_rate": round(pre_count / total_counted * 100) if total_counted else 0,
        "post_rate": round(post_count / total_counted * 100) if total_counted else 0,
        "kc_rate": round(kc_pass_count / total_counted * 100) if total_counted else 0,
        "all_three_rate": round(all_three / total_counted * 100) if total_counted else 0,
        "partners": partners,
        "workshop_pre_avg": workshop_pre_avg,
        "workshop_post_avg": workshop_post_avg,
        "workshop_delta": workshop_delta,
        "per_session": [
            {"slug": s["slug"], "name": s["name"], "start_date": s["start_date"], "end_date": s["end_date"],
             "counted": sd["total_counted"], "all_three": sd["all_three"], "all_three_rate": sd["all_three_rate"],
             "pre_avg": sd["pre_likert_avg"], "post_avg": sd["post_likert_avg"], "delta": sd["overall_delta"],
             "kc_rate": sd["kc_rate"]}
            for s, sd in zip(sessions, session_data)
        ],
        "today": date.today().isoformat(),
    }


@app.route("/admin/workshops/<slug>/rollup")
def admin_workshop_rollup(slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    workshop = get_db().execute("SELECT * FROM workshops WHERE slug = ?", (slug,)).fetchone()
    if workshop is None:
        abort(404)
    rollup = _workshop_rollup_data(workshop["id"], workshop["name"], workshop["slug"])
    if rollup is None:
        return render_template("admin_workshop_rollup_empty.html", user=user, workshop=workshop)
    return render_template("admin_workshop_rollup.html", user=user, workshop=workshop, data=rollup)


# Per-workshop leaderboard checkpoints. Order matters — defines column order on the leaderboard view.
WORKSHOP_LEADERBOARD = {
    "ifp": [
        {"key": "pre_survey",      "label": "Pre-Workshop Survey",   "type": "survey", "kind": "pre"},
        {"key": "kickoff",         "label": "View Kickoff",          "type": "page",   "path": "/w/ifp/01-overview.html"},
        {"key": "lab_1",           "label": "Lab 1",                 "type": "page",   "path": "/w/ifp/06-exercise-1.html"},
        {"key": "post_gen_1",      "label": "Post-Gen Tasks",        "type": "page",   "path": "/w/ifp/07-generation.html"},
        {"key": "lab_2",           "label": "Lab 2",                 "type": "page",   "path": "/w/ifp/06-exercise-2.html"},
        {"key": "post_gen_2",      "label": "Post-Gen Tasks",        "type": "page",   "path": "/w/ifp/08-post-gen.html"},
        {"key": "lab_3",           "label": "Lab 3",                 "type": "page",   "path": "/w/ifp/06-exercise-3.html"},
        {"key": "post_gen_3",      "label": "Post-Gen Tasks",        "type": "page",   "path": "/w/ifp/09-error-logs.html"},
        {"key": "tenant_cleanup",  "label": "Tenant Cleanup",        "type": "page",   "path": "/w/ifp/15-workshop-wrap-up.html"},
        {"key": "post_survey",     "label": "Post-Workshop Survey",  "type": "survey", "kind": "post"},
        {"key": "knowledge_check", "label": "Knowledge Check",       "type": "kc"},
    ],
}


def _session_leaderboard(session_id, workshop_slug):
    """One row per active participant with a status entry per checkpoint defined in
    WORKSHOP_LEADERBOARD for this workshop. Status is 'done' / 'partial' / 'todo'."""
    checkpoints = WORKSHOP_LEADERBOARD.get(workshop_slug, [])
    db = get_db()
    participants = db.execute(
        "SELECT sp.user_id, sp.partner, sp.excluded, u.username "
        "FROM session_participants sp JOIN users u ON u.id = sp.user_id "
        "WHERE sp.session_id = ? AND sp.removed_at IS NULL "
        "ORDER BY sp.partner IS NULL, sp.partner, u.username",
        (session_id,),
    ).fetchall()

    page_paths = [c["path"] for c in checkpoints if c["type"] == "page"]

    rows = []
    for p in participants:
        if page_paths:
            ph = ",".join("?" * len(page_paths))
            visited = {r["path"] for r in db.execute(
                f"SELECT DISTINCT path FROM page_views WHERE user_id = ? AND path IN ({ph})",
                [p["user_id"]] + page_paths,
            ).fetchall()}
        else:
            visited = set()

        pre_done = db.execute(
            "SELECT 1 FROM survey_responses sr JOIN session_surveys ss ON ss.id = sr.survey_id "
            "WHERE ss.session_id = ? AND ss.kind = 'pre' AND sr.participant_user_id = ? LIMIT 1",
            (session_id, p["user_id"]),
        ).fetchone() is not None
        post_done = db.execute(
            "SELECT 1 FROM survey_responses sr JOIN session_surveys ss ON ss.id = sr.survey_id "
            "WHERE ss.session_id = ? AND ss.kind = 'post' AND sr.participant_user_id = ? LIMIT 1",
            (session_id, p["user_id"]),
        ).fetchone() is not None
        kc_attempts, kc_passed = _kc_user_status(session_id, p["user_id"])

        statuses = []
        for c in checkpoints:
            if c["type"] == "survey":
                done = (c["kind"] == "pre" and pre_done) or (c["kind"] == "post" and post_done)
                statuses.append("done" if done else "todo")
            elif c["type"] == "page":
                statuses.append("done" if c["path"] in visited else "todo")
            elif c["type"] == "kc":
                if kc_passed:
                    statuses.append("done")
                elif kc_attempts:
                    statuses.append("partial")
                else:
                    statuses.append("todo")
            else:
                statuses.append("todo")

        # Display name: parse "First Last" from username if username looks like an email,
        # otherwise just use the username verbatim. Title-case the local part.
        display = p["username"]
        if "@" in display:
            local = display.split("@", 1)[0]
            parts = re.split(r"[._]", local)
            display = " ".join(p.title() for p in parts if p)

        rows.append({
            "user_id": p["user_id"],
            "username": p["username"],
            "display_name": display,
            "partner": p["partner"] or "Unknown",
            "excluded": bool(p["excluded"]),
            "statuses": statuses,
        })
    return rows, checkpoints


def _session_dashboard_rows(session_id, workshop_slug):
    """Per-participant progression snapshot for the live dashboard. One DB pass
    each for participants, then per-user queries for the small data — n usually
    < 100 so the per-user pattern stays cheap."""
    db = get_db()
    pages_meta = _list_workshop_pages(workshop_slug) or []
    page_paths = [f"/w/{workshop_slug}/{p['filename']}" for p in pages_meta]
    path_to_index = {p: i + 1 for i, p in enumerate(page_paths)}
    total_pages = len(page_paths)

    participants = db.execute(
        "SELECT sp.user_id, sp.partner, sp.excluded, u.username, u.last_login_at "
        "FROM session_participants sp JOIN users u ON u.id = sp.user_id "
        "WHERE sp.session_id = ? AND sp.removed_at IS NULL "
        "ORDER BY sp.partner IS NULL, sp.partner, u.username",
        (session_id,),
    ).fetchall()

    rows = []
    for p in participants:
        # Pages visited (limited to this workshop's content pages)
        visited = {r["path"] for r in db.execute(
            "SELECT DISTINCT path FROM page_views WHERE user_id = ? AND path LIKE ?",
            (p["user_id"], f"/w/{workshop_slug}/%.html"),
        ).fetchall()}
        visited_in_workshop = visited & set(page_paths)
        furthest_idx = max((path_to_index[v] for v in visited_in_workshop), default=0)
        furthest_name = pages_meta[furthest_idx - 1]["filename"] if furthest_idx else None

        last_row = db.execute(
            "SELECT MAX(viewed_at) AS last FROM page_views WHERE user_id = ?",
            (p["user_id"],),
        ).fetchone()
        last_activity = last_row["last"]

        # Survey completions for THIS session
        pre_done = db.execute(
            "SELECT 1 FROM survey_responses sr JOIN session_surveys ss ON ss.id = sr.survey_id "
            "WHERE ss.session_id = ? AND ss.kind = 'pre' AND sr.participant_user_id = ? LIMIT 1",
            (session_id, p["user_id"]),
        ).fetchone() is not None
        post_done = db.execute(
            "SELECT 1 FROM survey_responses sr JOIN session_surveys ss ON ss.id = sr.survey_id "
            "WHERE ss.session_id = ? AND ss.kind = 'post' AND sr.participant_user_id = ? LIMIT 1",
            (session_id, p["user_id"]),
        ).fetchone() is not None

        kc_attempts, kc_passed = _kc_user_status(session_id, p["user_id"])
        kc_best_row = db.execute(
            "SELECT MAX(sr.score) AS best FROM survey_responses sr "
            "JOIN session_surveys ss ON ss.id = sr.survey_id "
            "WHERE ss.session_id = ? AND ss.kind = 'knowledge' AND sr.participant_user_id = ?",
            (session_id, p["user_id"]),
        ).fetchone()
        kc_best_score = kc_best_row["best"] if kc_best_row else None

        rows.append({
            "user_id": p["user_id"],
            "username": p["username"],
            "partner": p["partner"] or "Unknown",
            "excluded": bool(p["excluded"]),
            "last_activity": last_activity,
            "last_login_at": p["last_login_at"],
            "visited_count": len(visited_in_workshop),
            "total_pages": total_pages,
            "furthest_idx": furthest_idx,
            "furthest_name": furthest_name,
            "pre_done": pre_done,
            "post_done": post_done,
            "kc_attempts": kc_attempts,
            "kc_best_score": kc_best_score,
            "kc_passed": kc_passed,
        })
    return rows, total_pages, pages_meta


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/leaderboard")
def admin_session_leaderboard(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    rows, checkpoints = _session_leaderboard(sess["id"], slug)
    return render_template(
        "admin_session_leaderboard.html",
        user=user, session=sess, rows=rows, checkpoints=checkpoints,
    )


# Backward-compat redirect for the old /dashboard URL
@app.route("/admin/workshops/<slug>/sessions/<session_slug>/dashboard")
def admin_session_dashboard(slug, session_slug):
    return redirect(url_for("admin_session_leaderboard", slug=slug, session_slug=session_slug))


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/reports")
def admin_session_reports(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    data = _session_report_data(sess["id"])
    return render_template("admin_reports.html", user=user, data=data, session=sess)


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/reports/exec-summary")
def admin_report_exec(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    data = _session_report_data(sess["id"])
    return render_template("report_exec_wrapper.html", user=user, data=data, session=sess)


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/reports/slack-post")
def admin_report_slack(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    data = _session_report_data(sess["id"])
    standout_names = ", ".join([b["name"] for b in data["standouts"][:5]]) if data["standouts"] else "[name 3-5 partner firms that stood out]"
    top3 = data["partners"][:3]
    medals = ["🥇", "🥈", "🥉"]
    top3_lines = "\n".join(
        f"{medals[i]} {b['name']} — {b['all_three']}/{b['count']} all three milestones ({round(b['rate']*100)}%)"
        for i, b in enumerate(top3)
    )
    next_open_date = "[DATE]"
    biggest = f'"{data["biggest_gain"]["header"]}"' if data["biggest_gain"] else '[TOPIC area]'
    net = ("+" if data["overall_delta"] is not None and data["overall_delta"] >= 0 else "") + (f"{data['overall_delta']}" if data["overall_delta"] is not None else "[X.XX]")
    pre_str = f"{data['pre_likert_avg']}" if data["pre_likert_avg"] is not None else "[X.XX]"
    post_str = f"{data['post_likert_avg']}" if data["post_likert_avg"] is not None else "[X.XX]"
    slack_text = f"""🎉 IFP Delivery Workshop — {sess['name']} just wrapped!

{data['partner_count']} delivery partners sent {data['total_counted']} consultants through the workshop this week. Quick highlights:

📈 Learning progression
• Pre-workshop survey average: {pre_str}
• Post-workshop survey average: {post_str}
• Net improvement: {net} points — biggest gains in {biggest}

🏆 Top completion rates by partner
{top3_lines}

✅ Knowledge check pass rate: {data['kc_rate']}%
🎯 {data['all_three_rate']}% of all participants completed all three milestones (pre-survey + post-survey + passed the knowledge check).

Huge shout-out to {standout_names} and to every consultant who put in the time. This is the kind of preparation that turns into smoother customer engagements.

🙏 Thanks also to our IFP product team who staffed the Slack channel all week — [count] questions answered, zero left hanging.

Next cohort opens {next_open_date}. If you have delivery consultants on your team who'd benefit, get them on the list — every session so far has hit capacity, and we're not extending tenant access past the registered week."""
    return render_template("report_slack.html", user=user, slack_text=slack_text, data=data, session=sess)


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/reports/partner-sponsor")
def admin_report_partner_landing(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    data = _session_report_data(sess["id"])
    return render_template("report_partner_landing.html", user=user, data=data, session=sess)


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/reports/partner-sponsor.zip")
def admin_report_partner_zip(slug, session_slug):
    import io
    import zipfile
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    data = _session_report_data(sess["id"])
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for partner in data["partners"]:
            inner_html = render_template(
                "report_partner_inner.html",
                data=data, partner=partner, session=sess,
            )
            full = (
                f"<!DOCTYPE html><html><head><meta charset=\"utf-8\">"
                f"<title>{partner['name']} — Partner Report</title></head>"
                f"<body style=\"background:#f8fafc;margin:0;padding:24px;\">{inner_html}</body></html>"
            )
            zf.writestr(f"{slug}-{session_slug}-{partner['slug']}-partner-report.html", full)
    buf.seek(0)
    from flask import send_file
    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"{slug}-{session_slug}-partner-reports.zip",
    )


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/reports/partner-sponsor/<partner_slug>")
def admin_report_partner(slug, session_slug, partner_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    data = _session_report_data(sess["id"])
    partner = next((b for b in data["partners"] if b["slug"] == partner_slug), None)
    if partner is None:
        abort(404)
    return render_template("report_partner_wrapper.html", user=user, data=data, partner=partner, session=sess)


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/reports/export.json")
def admin_report_export_json(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    data = _session_report_data(sess["id"])
    # Add raw response data for archival
    surveys = _load_session_surveys(sess["id"])
    out = {
        "session": data["session"],
        "summary": {k: data[k] for k in [
            "total_registered", "total_counted", "total_excluded",
            "pre_count", "post_count", "kc_pass_count", "all_three",
            "pre_rate", "post_rate", "kc_rate", "all_three_rate",
            "pre_likert_avg", "post_likert_avg", "overall_delta",
        ]},
        "partners": [{k: v for k, v in p.items() if k != "people"} for p in data["partners"]],
        "matched_likert": data["matched_likert"],
        "responses": {
            kind: [
                {
                    "user_id": r["user_id"],
                    "raw_email": r["raw_email"],
                    "responses": r["responses"],
                }
                for r in _load_survey_responses(surveys[kind]["id"])
            ]
            for kind in surveys
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    resp = jsonify(out)
    resp.headers["Content-Disposition"] = f"attachment; filename={sess['workshop_slug']}-{sess['slug']}-export.json"
    return resp


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/surveys", methods=["GET"])
def admin_session_surveys(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    surveys = _load_session_surveys(sess["id"])
    sections = []
    for kind in ("pre", "post", "knowledge"):
        s = surveys.get(kind)
        if not s:
            sections.append({"kind": kind, "survey": None})
            continue
        responses = _load_survey_responses(s["id"])
        stored_headers = json.loads(s["headers_json"]) if s["headers_json"] else []
        # Union of (a) stored headers (from XLSX uploads, legacy) and (b) keys
        # present on the response JSONs (native form submissions). Preserves
        # order from the stored list first, then appends any new ones seen.
        seen = set(stored_headers)
        headers = list(stored_headers)
        for r in responses:
            for k in r["responses"].keys():
                if k not in seen:
                    seen.add(k)
                    headers.append(k)
        cols = _classify_columns(headers, responses)
        aggregates = []
        for c in cols:
            agg = _aggregate_question(c["header"], c["qtype"], responses)
            agg["header"] = c["header"]
            aggregates.append(agg)
        sections.append({
            "kind": kind, "survey": s, "responses_count": len(responses),
            "is_anonymous": bool(s["is_anonymous"]), "aggregates": aggregates,
        })
    # Pre vs post Likert matching
    pre_aggs = next((s["aggregates"] for s in sections if s["kind"] == "pre" and s.get("aggregates")), [])
    post_aggs = next((s["aggregates"] for s in sections if s["kind"] == "post" and s.get("aggregates")), [])
    pre_by_h = {a["header"]: a for a in pre_aggs if a["type"] == "likert"}
    matched_likert = []
    for a in post_aggs:
        if a["type"] != "likert":
            continue
        pre = pre_by_h.get(a["header"])
        if not pre:
            continue
        delta = (a["avg"] - pre["avg"]) if (a["avg"] is not None and pre["avg"] is not None) else None
        matched_likert.append({
            "header": a["header"], "pre_avg": pre["avg"], "post_avg": a["avg"],
            "delta": round(delta, 2) if delta is not None else None,
            "pre_n": pre["n"], "post_n": a["n"],
        })
    return render_template(
        "admin_session_surveys.html",
        user=user, session=sess, sections=sections, matched_likert=matched_likert,
    )


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/surveys/upload", methods=["POST"])
def admin_session_survey_upload(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    kind = (request.form.get("kind") or "").strip().lower()
    if kind not in ("pre", "post", "knowledge"):
        abort(400)
    file = request.files.get("file")
    if not file or not file.filename:
        return redirect(url_for("admin_session_surveys", slug=slug, session_slug=session_slug))
    name_lower = file.filename.lower()
    try:
        if name_lower.endswith(".csv"):
            headers, rows = _parse_csv_upload(file)
        else:
            headers, rows = _parse_xlsx_upload(file)
    except Exception as e:
        return f"Failed to parse: {e}", 400
    _store_survey_upload(sess["id"], kind, file.filename, user["id"], headers, rows)
    return redirect(url_for("admin_session_surveys", slug=slug, session_slug=session_slug))


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/surveys/<kind>/remove", methods=["POST"])
def admin_session_survey_remove(slug, session_slug, kind):
    user, redir = _require_facilitator()
    if redir:
        return redir
    if kind not in ("pre", "post", "knowledge"):
        abort(400)
    sess = _get_session_or_404(slug, session_slug)
    get_db().execute("DELETE FROM session_surveys WHERE session_id = ? AND kind = ?", (sess["id"], kind))
    get_db().commit()
    return redirect(url_for("admin_session_surveys", slug=slug, session_slug=session_slug))


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/participants/<int:user_id>/remove", methods=["POST"])
def admin_session_participant_remove(slug, session_slug, user_id):
    cur_user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    db = get_db()
    db.execute(
        "UPDATE session_participants SET removed_at = CURRENT_TIMESTAMP "
        "WHERE session_id = ? AND user_id = ? AND removed_at IS NULL",
        (sess["id"], user_id),
    )
    db.commit()
    # Update the cached counts on each session_surveys row so the Collection
    # status card on the surveys page reflects the new totals immediately.
    survey_rows = db.execute(
        "SELECT id FROM session_surveys WHERE session_id = ?", (sess["id"],)
    ).fetchall()
    for r in survey_rows:
        _recompute_survey_counters(r["id"])
    return redirect(url_for("admin_session_participants", slug=slug, session_slug=session_slug))


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/monitors", methods=["GET", "POST"])
def admin_session_monitors(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    added_results = []
    if request.method == "POST":
        emails = _parse_email_list(request.form.get("emails", ""))
        if emails:
            added_results = _assign_monitors_to_session(sess["id"], emails, user["id"])
    monitors = _session_monitors(sess["id"])
    reset_creds = session.pop("reset_credentials", None)
    return render_template(
        "admin_session_monitors.html",
        user=user, session=sess, monitors=monitors, added_results=added_results,
        reset_creds=reset_creds,
    )


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/monitors/<int:monitor_id>/remove", methods=["POST"])
def admin_session_monitor_remove(slug, session_slug, monitor_id):
    cur_user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    get_db().execute(
        "UPDATE session_monitors SET removed_at = CURRENT_TIMESTAMP "
        "WHERE id = ? AND session_id = ? AND removed_at IS NULL",
        (monitor_id, sess["id"]),
    )
    get_db().commit()
    return redirect(url_for("admin_session_monitors", slug=slug, session_slug=session_slug))


def _reset_user_password(user_id):
    """Generate a new password, store the bcrypt hash, return the plaintext for the admin to forward."""
    new_pw = secrets.token_urlsafe(10)
    pw_hash = bcrypt.hashpw(new_pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    get_db().execute("UPDATE users SET password_hash = ? WHERE id = ?", (pw_hash, user_id))
    get_db().commit()
    return new_pw


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/monitors/<int:monitor_id>/reset-password", methods=["POST"])
def admin_session_monitor_reset_password(slug, session_slug, monitor_id):
    cur_user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    row = get_db().execute(
        "SELECT u.id, u.username FROM session_monitors sm JOIN users u ON u.id = sm.user_id "
        "WHERE sm.id = ? AND sm.session_id = ?",
        (monitor_id, sess["id"]),
    ).fetchone()
    if row is None:
        abort(404)
    new_pw = _reset_user_password(row["id"])
    session["reset_credentials"] = {"username": row["username"], "password": new_pw}
    return redirect(url_for("admin_session_monitors", slug=slug, session_slug=session_slug))


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/participants/<int:user_id>/reset-password", methods=["POST"])
def admin_session_participant_reset_password(slug, session_slug, user_id):
    cur_user, redir = _require_facilitator()
    if redir:
        return redir
    sess = _get_session_or_404(slug, session_slug)
    row = get_db().execute(
        "SELECT u.id, u.username FROM session_participants sp JOIN users u ON u.id = sp.user_id "
        "WHERE sp.user_id = ? AND sp.session_id = ? AND sp.removed_at IS NULL",
        (user_id, sess["id"]),
    ).fetchone()
    if row is None:
        abort(404)
    new_pw = _reset_user_password(row["id"])
    session["reset_credentials"] = {"username": row["username"], "password": new_pw}
    return redirect(url_for("admin_session_participants", slug=slug, session_slug=session_slug))


@app.route("/admin/workshops/<slug>/sessions/<session_slug>/archive", methods=["POST"])
def admin_session_archive(slug, session_slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    workshop = get_db().execute("SELECT id FROM workshops WHERE slug = ?", (slug,)).fetchone()
    if workshop is None:
        abort(404)
    cur = get_db().execute(
        "UPDATE sessions SET status = 'archived', archived_at = CURRENT_TIMESTAMP "
        "WHERE workshop_id = ? AND slug = ? AND status != 'archived'",
        (workshop["id"], session_slug),
    )
    get_db().commit()
    if cur.rowcount == 0:
        abort(404)
    return redirect(url_for("admin_workshop_detail", slug=slug))


@app.route("/admin/workshops/<slug>/archive", methods=["POST"])
def admin_workshop_archive(slug):
    user, redir = _require_facilitator()
    if redir:
        return redir
    cur = get_db().execute(
        "UPDATE workshops SET status = 'archived', archived_at = CURRENT_TIMESTAMP WHERE slug = ? AND status != 'archived'",
        (slug,),
    )
    get_db().commit()
    if cur.rowcount == 0:
        abort(404)
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/users-screen")
def admin_users_screen():
    user, redir = _require_facilitator()
    if redir:
        return redir
    users = get_db().execute(
        "SELECT id, username, display_name, email, is_facilitator, is_active, created_at, last_login_at "
        "FROM users ORDER BY created_at DESC"
    ).fetchall()
    return render_template("admin.html", user=user, users=users)


@app.route("/admin/users", methods=["POST"])
def admin_create_user():
    user = current_user()
    if user is None or not user["is_facilitator"]:
        abort(403)
    username = (request.form.get("username") or "").strip().lower()
    if not re.fullmatch(r"[a-z0-9._-]{3,40}", username):
        return jsonify(error="username must be 3-40 chars, lowercase alphanumeric + . _ -"), 400
    display_name = (request.form.get("display_name") or "").strip() or None
    email = (request.form.get("email") or "").strip() or None
    is_facilitator = 1 if request.form.get("is_facilitator") == "on" else 0
    initial_password = secrets.token_urlsafe(12)
    pw_hash = bcrypt.hashpw(initial_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    try:
        get_db().execute(
            "INSERT INTO users (username, password_hash, email, display_name, is_facilitator) VALUES (?, ?, ?, ?, ?)",
            (username, pw_hash, email, display_name, is_facilitator),
        )
        get_db().commit()
    except sqlite3.IntegrityError:
        return jsonify(error=f"username '{username}' already exists"), 409
    return jsonify(username=username, initial_password=initial_password, is_facilitator=bool(is_facilitator))


@app.route("/admin/users/<int:user_id>/reset-password", methods=["POST"])
def admin_reset_password(user_id):
    user = current_user()
    if user is None or not user["is_facilitator"]:
        abort(403)
    new_pw = secrets.token_urlsafe(12)
    pw_hash = bcrypt.hashpw(new_pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    cur = get_db().execute("UPDATE users SET password_hash = ? WHERE id = ?", (pw_hash, user_id))
    get_db().commit()
    if cur.rowcount == 0:
        return jsonify(error="user not found"), 404
    return jsonify(user_id=user_id, new_password=new_pw)


@app.route("/admin/users/<int:user_id>/toggle-active", methods=["POST"])
def admin_toggle_active(user_id):
    user = current_user()
    if user is None or not user["is_facilitator"]:
        abort(403)
    cur = get_db().execute("UPDATE users SET is_active = 1 - is_active WHERE id = ?", (user_id,))
    get_db().commit()
    if cur.rowcount == 0:
        return jsonify(error="user not found"), 404
    return jsonify(user_id=user_id)


@app.route("/admin/backup.db")
def admin_backup_db():
    """Stream a vacuumed snapshot of the SQLite DB. WAL-safe — uses .backup()."""
    import io
    user, redir = _require_facilitator()
    if redir:
        return redir
    src = get_db()
    buf_path = Path("/tmp") / f"workshop-backup-{int(time.time())}.db"
    try:
        dest = sqlite3.connect(str(buf_path))
        with dest:
            src.backup(dest)
        dest.close()
        data = buf_path.read_bytes()
    finally:
        try:
            buf_path.unlink()
        except OSError:
            pass
    from flask import send_file
    return send_file(
        io.BytesIO(data),
        mimetype="application/x-sqlite3",
        as_attachment=True,
        download_name=f"workshop-backup-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.db",
    )


@app.route("/admin/views/<int:user_id>")
def admin_user_views(user_id):
    user = current_user()
    if user is None or not user["is_facilitator"]:
        abort(403)
    target = get_db().execute("SELECT id, username, display_name FROM users WHERE id = ?", (user_id,)).fetchone()
    if target is None:
        abort(404)
    views = get_db().execute(
        "SELECT path, status, viewed_at FROM page_views WHERE user_id = ? ORDER BY viewed_at DESC LIMIT 200",
        (user_id,),
    ).fetchall()
    return render_template("user_views.html", user=user, target=target, views=views)


# ---------------------------- CLI: bootstrap admin ----------------------------

@app.cli.command("init-db")
def cli_init_db():
    """Create the SQLite tables."""
    init_db()
    print(f"initialized {DB_PATH}")


@app.cli.command("add-admin")
def cli_add_admin():
    """Create the first facilitator account interactively."""
    import getpass
    init_db()
    username = input("Username: ").strip().lower()
    if not re.fullmatch(r"[a-z0-9._-]{3,40}", username):
        raise SystemExit("invalid username")
    display = input("Display name (optional): ").strip() or None
    email = input("Email (optional): ").strip() or None
    pw1 = getpass.getpass("Password: ")
    pw2 = getpass.getpass("Confirm: ")
    if pw1 != pw2 or len(pw1) < 8:
        raise SystemExit("passwords don't match or under 8 chars")
    pw_hash = bcrypt.hashpw(pw1.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    conn = sqlite3.connect(str(DB_PATH))
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, email, display_name, is_facilitator) VALUES (?, ?, ?, ?, 1)",
            (username, pw_hash, email, display),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        raise SystemExit(f"username '{username}' already exists")
    finally:
        conn.close()
    print(f"created facilitator: {username}")


# ---------------------------- Discussion Q&A panel (B/20) ----------------------------

QA_PAGE_RE = re.compile(r"^/w/([a-z0-9-]+)/([a-z0-9_-]+)\.html$")


def _display_name_from_username(username):
    if not username or "@" not in username:
        return username or "Unknown"
    local = username.split("@", 1)[0]
    parts = re.split(r"[._-]+", local)
    return " ".join(p.capitalize() for p in parts if p) or username


def _user_workshop_session(user_id, workshop_slug):
    """Return (session_row, role) for this user in this workshop.
    role is 'participant', 'monitor', or both None when not enrolled and not assigned."""
    if user_id is None:
        return (None, None)
    sess = _current_enrollment_session(user_id, workshop_slug)
    if sess:
        return (sess, "participant")
    sess = get_db().execute(
        "SELECT s.id, s.slug, s.name FROM session_monitors sm "
        "JOIN sessions s ON s.id = sm.session_id "
        "JOIN workshops w ON w.id = s.workshop_id "
        "WHERE sm.user_id = ? AND sm.removed_at IS NULL "
        "  AND w.slug = ? AND s.status != 'archived' "
        "ORDER BY sm.assigned_at DESC LIMIT 1",
        (user_id, workshop_slug),
    ).fetchone()
    if sess:
        return (sess, "monitor")
    return (None, None)


def _session_has_any_monitor(session_id):
    return get_db().execute(
        "SELECT 1 FROM session_monitors WHERE session_id = ? AND removed_at IS NULL LIMIT 1",
        (session_id,),
    ).fetchone() is not None


def _accessible_sessions_for_workshop(user_id, workshop_slug):
    """Sessions the user can view in this workshop, most-recent first.
    Facilitators see all non-archived sessions; others see only the ones
    they are enrolled in or assigned to as a monitor."""
    if user_id is None:
        return []
    db = get_db()
    user_row = db.execute("SELECT is_facilitator FROM users WHERE id = ?", (user_id,)).fetchone()
    if user_row and user_row["is_facilitator"]:
        return db.execute(
            "SELECT s.id, s.slug, s.name FROM sessions s "
            "JOIN workshops w ON w.id = s.workshop_id "
            "WHERE w.slug = ? AND s.status != 'archived' "
            "ORDER BY s.start_date DESC",
            (workshop_slug,),
        ).fetchall()
    return db.execute(
        "SELECT DISTINCT s.id, s.slug, s.name FROM sessions s "
        "JOIN workshops w ON w.id = s.workshop_id "
        "LEFT JOIN session_participants sp ON sp.session_id = s.id AND sp.user_id = ? AND sp.removed_at IS NULL "
        "LEFT JOIN session_monitors sm ON sm.session_id = s.id AND sm.user_id = ? AND sm.removed_at IS NULL "
        "WHERE w.slug = ? AND s.status != 'archived' "
        "  AND (sp.id IS NOT NULL OR sm.id IS NOT NULL) "
        "ORDER BY s.start_date DESC",
        (user_id, user_id, workshop_slug),
    ).fetchall()


def _threads_for_page(session_id, page_path, viewer_user_id):
    """Return list of threads on this (session, page) with reactions, replies, and viewer-specific flags."""
    db = get_db()
    threads_rows = db.execute(
        "SELECT t.id, t.author_user_id, t.anonymous, t.kind, t.body, t.status, t.created_at, t.answered_at, "
        "       u.username AS author_username, "
        "       (SELECT COUNT(*) FROM reactions r WHERE r.thread_id = t.id AND r.kind = 'me_too') AS me_too_count, "
        "       EXISTS(SELECT 1 FROM reactions r WHERE r.thread_id = t.id AND r.user_id = ? AND r.kind = 'me_too') AS i_reacted, "
        "       (SELECT MAX(rep.created_at) FROM replies rep WHERE rep.thread_id = t.id AND rep.hidden_at IS NULL) AS latest_reply_at, "
        "       (SELECT rs.last_read_at FROM read_state rs WHERE rs.thread_id = t.id AND rs.user_id = ?) AS last_read_at "
        "FROM threads t LEFT JOIN users u ON u.id = t.author_user_id "
        "WHERE t.session_id = ? AND t.page_path = ? AND t.hidden_at IS NULL "
        "ORDER BY t.created_at",
        (viewer_user_id, viewer_user_id, session_id, page_path),
    ).fetchall()
    if not threads_rows:
        return []
    thread_ids = [t["id"] for t in threads_rows]
    placeholders = ",".join("?" * len(thread_ids))
    replies_rows = db.execute(
        f"SELECT r.id, r.thread_id, r.author_user_id, r.body, r.is_monitor_reply, r.created_at, "
        f"       u.username AS author_username "
        f"FROM replies r LEFT JOIN users u ON u.id = r.author_user_id "
        f"WHERE r.thread_id IN ({placeholders}) AND r.hidden_at IS NULL "
        f"ORDER BY r.created_at",
        thread_ids,
    ).fetchall()
    by_thread = {}
    for r in replies_rows:
        d = dict(r)
        d["author_display"] = _display_name_from_username(r["author_username"])
        by_thread.setdefault(r["thread_id"], []).append(d)
    result = []
    for t in threads_rows:
        d = dict(t)
        d["author_display"] = "Anonymous" if t["anonymous"] else _display_name_from_username(t["author_username"])
        d["replies"] = by_thread.get(t["id"], [])
        latest_activity = max(t["created_at"], t["latest_reply_at"] or "")
        d["new_for_me"] = (t["last_read_at"] is None) or (latest_activity > t["last_read_at"])
        result.append(d)
    return result


def _mark_threads_read(user_id, thread_ids):
    """Upsert read_state to now for the given threads."""
    if not thread_ids or user_id is None:
        return
    db = get_db()
    db.executemany(
        "INSERT INTO read_state (user_id, thread_id, last_read_at) VALUES (?, ?, CURRENT_TIMESTAMP) "
        "ON CONFLICT(user_id, thread_id) DO UPDATE SET last_read_at = excluded.last_read_at",
        [(user_id, tid) for tid in thread_ids],
    )
    db.commit()


def _page_label_from_path(path):
    """Pretty label for /w/ifp/01-overview.html → '01 Overview' (cheap, no template parse)."""
    m = re.match(r"^/w/[a-z0-9-]+/(.+)\.html$", path)
    if not m:
        return path
    parts = m.group(1).split("-")
    return " ".join(p.capitalize() if not p.isdigit() else p for p in parts)


def _cohort_feed(session_id, viewer_user_id, filter_kind="all"):
    """Return threads across all pages in this cohort, newest first, with viewer-specific flags."""
    db = get_db()
    where_extra = ""
    if filter_kind == "unanswered":
        where_extra = " AND t.status = 'open' AND t.kind = 'question'"
    elif filter_kind == "celebrations":
        where_extra = " AND t.kind = 'celebration'"
    threads_rows = db.execute(
        "SELECT t.id, t.author_user_id, t.anonymous, t.kind, t.body, t.status, t.created_at, t.page_path, "
        "       u.username AS author_username, "
        "       (SELECT COUNT(*) FROM reactions r WHERE r.thread_id = t.id AND r.kind = 'me_too') AS me_too_count, "
        "       (SELECT COUNT(*) FROM replies rep WHERE rep.thread_id = t.id AND rep.hidden_at IS NULL) AS reply_count, "
        "       (SELECT MAX(rep.created_at) FROM replies rep WHERE rep.thread_id = t.id AND rep.hidden_at IS NULL) AS latest_reply_at, "
        "       EXISTS(SELECT 1 FROM reactions r WHERE r.thread_id = t.id AND r.user_id = ? AND r.kind = 'me_too') AS i_reacted, "
        "       (SELECT rs.last_read_at FROM read_state rs WHERE rs.thread_id = t.id AND rs.user_id = ?) AS last_read_at "
        "FROM threads t LEFT JOIN users u ON u.id = t.author_user_id "
        f"WHERE t.session_id = ? AND t.hidden_at IS NULL{where_extra} "
        "ORDER BY t.created_at DESC "
        "LIMIT 200",
        (viewer_user_id, viewer_user_id, session_id),
    ).fetchall()
    results = []
    for t in threads_rows:
        d = dict(t)
        d["author_display"] = "Anonymous" if t["anonymous"] else _display_name_from_username(t["author_username"])
        latest_activity = max(t["created_at"], t["latest_reply_at"] or "")
        d["new_for_me"] = (t["last_read_at"] is None) or (latest_activity > t["last_read_at"])
        d["page_label"] = _page_label_from_path(t["page_path"])
        results.append(d)
    return results


def _default_landing_for_user(user_id):
    """Return URL of the most-recent landing for this user, or None.
    Participants land on the cohort feed; pure monitors land on the monitor dashboard."""
    if user_id is None:
        return None
    db = get_db()
    row = db.execute(
        "SELECT s.slug AS session_slug, w.slug AS workshop_slug FROM session_participants sp "
        "JOIN sessions s ON s.id = sp.session_id "
        "JOIN workshops w ON w.id = s.workshop_id "
        "WHERE sp.user_id = ? AND sp.removed_at IS NULL AND s.status != 'archived' "
        "ORDER BY sp.enrolled_at DESC LIMIT 1",
        (user_id,),
    ).fetchone()
    if row:
        return f"/w/{row['workshop_slug']}/s/{row['session_slug']}/feed"
    row = db.execute(
        "SELECT s.slug AS session_slug, w.slug AS workshop_slug FROM session_monitors sm "
        "JOIN sessions s ON s.id = sm.session_id "
        "JOIN workshops w ON w.id = s.workshop_id "
        "WHERE sm.user_id = ? AND sm.removed_at IS NULL AND s.status != 'archived' "
        "ORDER BY sm.assigned_at DESC LIMIT 1",
        (user_id,),
    ).fetchone()
    if row:
        return f"/m/{row['workshop_slug']}/{row['session_slug']}"
    return None


def _monitor_dashboard_open_threads(session_id):
    """Open question threads in this session, oldest-first, with replies and aging flag."""
    db = get_db()
    threads_rows = db.execute(
        "SELECT t.id, t.author_user_id, t.anonymous, t.kind, t.body, t.status, t.created_at, t.page_path, "
        "       u.username AS author_username, "
        "       (SELECT COUNT(*) FROM reactions r WHERE r.thread_id = t.id AND r.kind = 'me_too') AS me_too_count "
        "FROM threads t LEFT JOIN users u ON u.id = t.author_user_id "
        "WHERE t.session_id = ? AND t.status = 'open' AND t.kind = 'question' AND t.hidden_at IS NULL "
        "ORDER BY t.created_at ASC LIMIT 50",
        (session_id,),
    ).fetchall()
    if not threads_rows:
        return []
    thread_ids = [t["id"] for t in threads_rows]
    placeholders = ",".join("?" * len(thread_ids))
    replies_rows = db.execute(
        f"SELECT r.id, r.thread_id, r.body, r.is_monitor_reply, r.created_at, u.username AS author_username "
        f"FROM replies r LEFT JOIN users u ON u.id = r.author_user_id "
        f"WHERE r.thread_id IN ({placeholders}) AND r.hidden_at IS NULL "
        f"ORDER BY r.created_at",
        thread_ids,
    ).fetchall()
    by_thread = {}
    for r in replies_rows:
        d = dict(r)
        d["author_display"] = _display_name_from_username(r["author_username"])
        by_thread.setdefault(r["thread_id"], []).append(d)
    now = datetime.utcnow()
    result = []
    for t in threads_rows:
        d = dict(t)
        d["author_display"] = "Anonymous" if t["anonymous"] else _display_name_from_username(t["author_username"])
        d["replies"] = by_thread.get(t["id"], [])
        d["page_label"] = _page_label_from_path(t["page_path"])
        try:
            created = datetime.fromisoformat(t["created_at"].replace(" ", "T"))
            age_seconds = (now - created).total_seconds()
            d["is_aging"] = age_seconds > 3600
            d["age_minutes"] = max(0, round(age_seconds / 60))
        except (ValueError, TypeError, AttributeError):
            d["is_aging"] = False
            d["age_minutes"] = None
        result.append(d)
    return result


def _monitor_dashboard_metrics(session_id, session_row):
    """Open count, median response time (minutes), engagement, days remaining."""
    db = get_db()
    open_count = db.execute(
        "SELECT COUNT(*) AS c FROM threads "
        "WHERE session_id = ? AND status = 'open' AND kind = 'question' AND hidden_at IS NULL",
        (session_id,),
    ).fetchone()["c"]
    response_times = db.execute(
        "SELECT (julianday(MIN(rep.created_at)) - julianday(t.created_at)) * 24 * 60 AS minutes "
        "FROM threads t JOIN replies rep ON rep.thread_id = t.id "
        "WHERE t.session_id = ? AND t.status = 'answered' AND t.kind = 'question' AND rep.is_monitor_reply = 1 "
        "GROUP BY t.id",
        (session_id,),
    ).fetchall()
    times_sorted = sorted(r["minutes"] for r in response_times if r["minutes"] is not None and r["minutes"] >= 0)
    median_minutes = None
    if times_sorted:
        mid = len(times_sorted) // 2
        median_minutes = times_sorted[mid] if len(times_sorted) % 2 == 1 else (times_sorted[mid - 1] + times_sorted[mid]) / 2
    total_participants = db.execute(
        "SELECT COUNT(*) AS c FROM session_participants WHERE session_id = ? AND removed_at IS NULL AND excluded = 0",
        (session_id,),
    ).fetchone()["c"]
    posted_count = db.execute(
        "SELECT COUNT(DISTINCT t.author_user_id) AS c FROM threads t "
        "JOIN session_participants sp ON sp.user_id = t.author_user_id AND sp.session_id = t.session_id "
        "WHERE t.session_id = ? AND sp.removed_at IS NULL AND sp.excluded = 0 AND t.hidden_at IS NULL",
        (session_id,),
    ).fetchone()["c"]
    days_remaining = None
    if session_row["end_date"]:
        try:
            end = date.fromisoformat(session_row["end_date"])
            days_remaining = (end - date.today()).days
        except (ValueError, TypeError):
            days_remaining = None
    return {
        "open_count": open_count,
        "median_response_minutes": median_minutes,
        "posted_count": posted_count,
        "total_participants": total_participants,
        "days_remaining": days_remaining,
    }


def _safe_return_to(value, fallback):
    """Validate a return_to form field to prevent open-redirect; fall back if invalid."""
    if not value:
        return fallback
    value = value.strip()
    if value.startswith("/") and not value.startswith("//"):
        return value
    return fallback


@app.after_request
def inject_qa_panel(resp):
    """Inject the discussion panel before </main> on workshop content pages
    when the viewer has an active session enrollment or monitor assignment."""
    if request.method != "GET":
        return resp
    if resp.status_code != 200:
        return resp
    m = QA_PAGE_RE.match(request.path)
    if not m:
        return resp
    user = g.get("user")
    if user is None:
        return resp
    content_type = resp.headers.get("Content-Type", "")
    if not content_type.startswith("text/html"):
        return resp
    workshop_slug = m.group(1)
    accessible_sessions = _accessible_sessions_for_workshop(user["id"], workshop_slug)
    if not accessible_sessions:
        return resp
    desired = request.args.get("session")
    sess = next((s for s in accessible_sessions if s["slug"] == desired), None) if desired else None
    if sess is None:
        sess = accessible_sessions[0]
    if _user_is_session_monitor(sess["id"], user["id"]):
        role = "monitor"
    elif user["is_facilitator"]:
        role = "facilitator"
    else:
        role = "participant"
    has_monitor = _session_has_any_monitor(sess["id"])
    threads = _threads_for_page(sess["id"], request.path, user["id"]) if has_monitor else []
    new_count = sum(1 for t in threads if t.get("new_for_me"))
    panel_html = render_template(
        "_qa_panel.html",
        threads=threads,
        session_name=sess["name"],
        session_slug=sess["slug"],
        workshop_slug=workshop_slug,
        page_path=request.path,
        role=role,
        has_monitor=has_monitor,
        new_count=new_count,
        accessible_sessions=accessible_sessions,
    )
    # send_from_directory returns the response in direct_passthrough mode,
    # which makes resp.get_data() raise. Switch off before reading the body.
    resp.direct_passthrough = False
    body = resp.get_data(as_text=True)
    if "</main>" in body:
        body = body.replace("</main>", panel_html + "\n  </main>", 1)
        resp.set_data(body)
        resp.headers["Content-Length"] = str(len(resp.get_data()))
    if threads:
        _mark_threads_read(user["id"], [t["id"] for t in threads])
    return resp


@app.route("/m/<workshop_slug>/<session_slug>")
def monitor_dashboard(workshop_slug, session_slug):
    user = current_user()
    if user is None:
        abort(403)
    sess = _get_session_or_404(workshop_slug, session_slug)
    is_monitor = _user_is_session_monitor(sess["id"], user["id"])
    if not (is_monitor or user["is_facilitator"]):
        abort(403)
    open_threads = _monitor_dashboard_open_threads(sess["id"])
    metrics = _monitor_dashboard_metrics(sess["id"], sess)
    return render_template(
        "monitor_dashboard.html",
        user=user, session=sess, workshop_slug=workshop_slug,
        open_threads=open_threads, metrics=metrics,
        is_monitor=is_monitor,
    )


@app.route("/w/<workshop_slug>/s/<session_slug>/feed")
def cohort_feed(workshop_slug, session_slug):
    user = current_user()
    if user is None:
        abort(403)
    sess = _get_session_or_404(workshop_slug, session_slug)
    is_participant = get_db().execute(
        "SELECT 1 FROM session_participants WHERE session_id = ? AND user_id = ? AND removed_at IS NULL",
        (sess["id"], user["id"]),
    ).fetchone() is not None
    is_monitor = _user_is_session_monitor(sess["id"], user["id"])
    if not (is_participant or is_monitor or user["is_facilitator"]):
        abort(403)
    filter_kind = request.args.get("filter", "all")
    if filter_kind not in ("all", "unanswered", "celebrations"):
        filter_kind = "all"
    threads = _cohort_feed(sess["id"], user["id"], filter_kind)
    if threads:
        _mark_threads_read(user["id"], [t["id"] for t in threads])
    role = "monitor" if is_monitor else ("facilitator" if user["is_facilitator"] else "participant")
    has_monitor = _session_has_any_monitor(sess["id"])
    return render_template(
        "cohort_feed.html",
        user=user, session=sess, threads=threads, filter_kind=filter_kind,
        role=role, has_monitor=has_monitor, workshop_slug=workshop_slug,
    )


@app.route("/w/<workshop_slug>/threads", methods=["POST"])
def qa_create_thread(workshop_slug):
    user = current_user()
    if user is None:
        abort(403)
    sess, role = _user_workshop_session(user["id"], workshop_slug)
    if sess is None or not _session_has_any_monitor(sess["id"]):
        abort(403)
    page_path = request.form.get("page_path", "").strip()
    body = (request.form.get("body") or "").strip()
    anonymous = 1 if request.form.get("anonymous") == "1" else 0
    if not page_path or not body or not QA_PAGE_RE.match(page_path):
        abort(400)
    cur = get_db().execute(
        "INSERT INTO threads (session_id, page_path, author_user_id, anonymous, kind, body) "
        "VALUES (?, ?, ?, ?, 'question', ?)",
        (sess["id"], page_path, user["id"], anonymous, body),
    )
    new_thread_id = cur.lastrowid
    get_db().commit()
    return redirect(page_path + f"#qa-thread-{new_thread_id}")


@app.route("/w/<workshop_slug>/threads/<int:thread_id>/reply", methods=["POST"])
def qa_create_reply(workshop_slug, thread_id):
    user = current_user()
    if user is None:
        abort(403)
    sess, role = _user_workshop_session(user["id"], workshop_slug)
    if sess is None:
        abort(403)
    thread = get_db().execute(
        "SELECT id, page_path FROM threads WHERE id = ? AND hidden_at IS NULL AND session_id = ?",
        (thread_id, sess["id"]),
    ).fetchone()
    if thread is None:
        abort(404)
    body = (request.form.get("body") or "").strip()
    if not body:
        abort(400)
    is_monitor_reply = 1 if _user_is_session_monitor(sess["id"], user["id"]) else 0
    get_db().execute(
        "INSERT INTO replies (thread_id, author_user_id, body, is_monitor_reply) "
        "VALUES (?, ?, ?, ?)",
        (thread_id, user["id"], body, is_monitor_reply),
    )
    get_db().commit()
    return redirect(_safe_return_to(request.form.get("return_to"), thread["page_path"] + f"#qa-thread-{thread_id}"))


@app.route("/w/<workshop_slug>/threads/<int:thread_id>/react", methods=["POST"])
def qa_react(workshop_slug, thread_id):
    user = current_user()
    if user is None:
        abort(403)
    sess, role = _user_workshop_session(user["id"], workshop_slug)
    if sess is None:
        abort(403)
    thread = get_db().execute(
        "SELECT id, page_path FROM threads WHERE id = ? AND hidden_at IS NULL AND session_id = ?",
        (thread_id, sess["id"]),
    ).fetchone()
    if thread is None:
        abort(404)
    existing = get_db().execute(
        "SELECT 1 FROM reactions WHERE thread_id = ? AND user_id = ? AND kind = 'me_too'",
        (thread_id, user["id"]),
    ).fetchone()
    if existing:
        get_db().execute(
            "DELETE FROM reactions WHERE thread_id = ? AND user_id = ? AND kind = 'me_too'",
            (thread_id, user["id"]),
        )
    else:
        get_db().execute(
            "INSERT INTO reactions (thread_id, user_id, kind) VALUES (?, ?, 'me_too')",
            (thread_id, user["id"]),
        )
    get_db().commit()
    return redirect(thread["page_path"] + f"#qa-thread-{thread_id}")


@app.route("/w/<workshop_slug>/threads/<int:thread_id>/answer", methods=["POST"])
def qa_mark_answered(workshop_slug, thread_id):
    user = current_user()
    if user is None:
        abort(403)
    sess, role = _user_workshop_session(user["id"], workshop_slug)
    if sess is None or not _user_is_session_monitor(sess["id"], user["id"]):
        abort(403)
    thread = get_db().execute(
        "SELECT id, page_path FROM threads WHERE id = ? AND hidden_at IS NULL AND session_id = ?",
        (thread_id, sess["id"]),
    ).fetchone()
    if thread is None:
        abort(404)
    get_db().execute(
        "UPDATE threads SET status = 'answered', answered_at = CURRENT_TIMESTAMP, answered_by = ? "
        "WHERE id = ?",
        (user["id"], thread_id),
    )
    get_db().commit()
    return redirect(_safe_return_to(request.form.get("return_to"), thread["page_path"] + f"#qa-thread-{thread_id}"))


# Initialize DB at import time so gunicorn workers see the schema.
init_db()
